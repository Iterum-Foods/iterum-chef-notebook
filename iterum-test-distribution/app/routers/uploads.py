from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime
import os
import uuid
from typing import Optional
import json

from app.database import get_db, RecipeUpload, User
from app.schemas import RecipeUpload as RecipeUploadSchema
from app.core.auth import get_current_user
from app.core.config import settings
from app.services.recipe_parser import RecipeParser
import openpyxl
from docx import Document
import io

router = APIRouter()


@router.post("/recipe", response_model=RecipeUploadSchema)
async def upload_recipe_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload a recipe file (image, PDF, etc.) for OCR processing"""
    
    # Validate file type
    allowed_types = [
        "image/jpeg", "image/png", "image/jpg", "application/pdf",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel",  # .xls
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
        "application/msword"  # .doc
    ]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail=f"File type {file.content_type} not supported. Allowed types: {allowed_types}"
        )
    
    # Validate file size
    if file.size and file.size > settings.MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File size {file.size} exceeds maximum allowed size of {settings.MAX_FILE_SIZE}"
        )
    
    # Generate unique filename
    file_extension = os.path.splitext(file.filename or "")[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(settings.UPLOAD_DIR, unique_filename)
    
    # Save file
    try:
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
    
    # Create upload record
    db_upload = RecipeUpload(
        filename=unique_filename,
        original_filename=file.filename,
        file_path=file_path,
        file_size=len(content),
        mime_type=file.content_type,
        upload_status="pending",
        uploaded_by=current_user.id
    )
    
    db.add(db_upload)
    db.commit()
    db.refresh(db_upload)
    
    # TODO: Start background task for OCR processing
    # This would be implemented with Celery or similar
    
    return db_upload


@router.get("/", response_model=list[RecipeUploadSchema])
async def get_uploads(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all uploads for the current user"""
    
    uploads = db.query(RecipeUpload).filter(
        RecipeUpload.uploaded_by == current_user.id
    ).order_by(RecipeUpload.created_at.desc()).all()
    
    return uploads


@router.get("/{upload_id}", response_model=RecipeUploadSchema)
async def get_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a specific upload by ID"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    return upload


@router.delete("/{upload_id}")
async def delete_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an upload and its associated file"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # Delete file from filesystem
    try:
        file_path_str = str(upload.file_path)
        if os.path.exists(file_path_str):
            os.remove(file_path_str)
    except Exception as e:
        # Log error but continue with database deletion
        print(f"Failed to delete file {upload.file_path}: {e}")
    
    # Delete from database
    db.delete(upload)
    db.commit()
    
    return {"message": "Upload deleted successfully"}


@router.post("/{upload_id}/process")
async def process_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Manually trigger OCR processing for an upload"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # Update status to processing
    upload.upload_status = str("processing")
    db.commit()
    
    # TODO: Implement actual OCR processing
    # This would use pytesseract or similar for image processing
    # For now, return a placeholder response
    
    return {
        "message": "OCR processing started",
        "upload_id": upload_id,
        "note": "OCR processing will be implemented in the next iteration"
    }


@router.get("/{upload_id}/ocr-text")
async def get_ocr_text(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get the OCR text from a processed upload"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if upload.upload_status != "completed":
        raise HTTPException(
            status_code=400,
            detail="Upload has not been processed yet"
        )
    
    return {
        "upload_id": upload_id,
        "ocr_text": upload.ocr_text,
        "parsed_data": upload.parsed_data
    }


@router.post("/parse-text")
async def parse_recipe_text(
    text_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Parse recipe text and extract structured information"""
    try:
        recipe_text = text_data.get("text", "")
        print(f"[DEBUG] Incoming recipe text: {recipe_text[:100]}...")
        if not recipe_text:
            raise HTTPException(status_code=400, detail="No text provided")
        parser = RecipeParser()
        parsed_recipe = parser.parse_recipe_text(recipe_text)
        result = parser.to_dict(parsed_recipe)
        print(f"[DEBUG] Parsed recipe result: {json.dumps(result)[:200]}...")
        return result
    except Exception as e:
        print(f"[ERROR] Failed to parse recipe: {e}")
        raise HTTPException(status_code=500, detail=f"Error parsing recipe: {str(e)}")

# --- Add /extract-text endpoint if missing ---
@router.post("/extract-text")
async def extract_text_from_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Extract text from uploaded file (PDF, Word, Excel, etc.)"""
    try:
        # For now, just return a stub response
        return {"message": "Text extraction not yet implemented. Please use /parse-text with raw text."}
    except Exception as e:
        print(f"[ERROR] Failed to extract text: {e}")
        raise HTTPException(status_code=500, detail=f"Error extracting text: {str(e)}")


@router.post("/test-parse")
async def test_recipe_parsing(
    recipe_text: str,
    current_user: User = Depends(get_current_user)
):
    """Test endpoint to parse recipe text and extract structured information"""
    
    # Initialize recipe parser
    parser = RecipeParser()
    
    try:
        # Parse the recipe text
        parsed_recipe = parser.parse_recipe_text(recipe_text)
        
        # Convert to dictionary format
        parsed_data = parser.to_dict(parsed_recipe)
        
        return {
            "message": "Recipe parsed successfully",
            "parsed_data": parsed_data,
            "summary": {
                "title": parsed_recipe.title,
                "ingredients_count": len(parsed_recipe.ingredients),
                "instructions_count": len(parsed_recipe.instructions),
                "allergy_notes_count": len(parsed_recipe.allergy_notes),
                "pricing_info": parsed_recipe.pricing_info,
                "yield_info": parsed_recipe.yield_info,
                "prep_time": parsed_recipe.prep_time,
                "cook_time": parsed_recipe.cook_time,
                "difficulty": parsed_recipe.difficulty,
                "cuisine": parsed_recipe.cuisine,
                "tags": parsed_recipe.tags
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse recipe: {str(e)}")


@router.post("/{upload_id}/parse")
async def parse_uploaded_recipe(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Parse uploaded recipe text and extract structured information"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if not upload.ocr_text:
        raise HTTPException(
            status_code=400,
            detail="No OCR text available. Please process the upload first."
        )
    
    # Initialize recipe parser
    parser = RecipeParser()
    
    try:
        # Parse the recipe text
        ocr_text = str(upload.ocr_text)
        parsed_recipe = parser.parse_recipe_text(ocr_text)
        
        # Convert to dictionary format
        parsed_data = parser.to_dict(parsed_recipe)
        
        # Update upload with parsed data
        upload.parsed_data = parsed_data
        upload.upload_status = "completed"
        db.commit()
        
        return {
            "message": "Recipe parsed successfully",
            "upload_id": upload_id,
            "parsed_data": parsed_data
        }
        
    except Exception as e:
        upload.error_message = str(e)
        upload.upload_status = "failed"
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to parse recipe: {str(e)}")


@router.get("/{upload_id}/parsed-data")
async def get_parsed_recipe_data(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get parsed recipe data from an upload"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if not upload.parsed_data:
        raise HTTPException(
            status_code=400,
            detail="No parsed data available. Please parse the upload first."
        )
    
    return {
        "upload_id": upload_id,
        "parsed_data": upload.parsed_data
    }


@router.post("/{upload_id}/create-recipe")
async def create_recipe_from_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a recipe from a processed upload"""
    
    upload = db.query(RecipeUpload).filter(
        RecipeUpload.id == upload_id,
        RecipeUpload.uploaded_by == current_user.id
    ).first()
    
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    if upload.upload_status != "completed":
        raise HTTPException(
            status_code=400,
            detail="Upload must be processed before creating a recipe"
        )
    
    # TODO: Implement recipe creation from parsed data
    # This would use the parsed_data to create a new recipe
    
    return {
        "message": "Recipe creation from upload will be implemented",
        "upload_id": upload_id,
        "note": "This will create a recipe from the OCR parsed data"
    }


@router.post("/extract-text")
async def extract_text_from_document(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Extract text from Excel or Word documents"""
    try:
        content = await file.read()
        
        # Handle Excel files (.xlsx, .xls)
        if file.content_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ]:
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                text_content = []
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text_content.append(f"=== Sheet: {sheet_name} ===")
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        if row_text.strip():
                            text_content.append(row_text)
                    
                    text_content.append("")  # Empty line between sheets
                
                extracted_text = "\n".join(text_content)
                
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to read Excel file: {str(e)}"
                )
        
        # Handle Word files (.docx, .doc)
        elif file.content_type in [
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword"
        ]:
            try:
                # Note: python-docx only supports .docx, not .doc
                if file.content_type == "application/msword":
                    raise HTTPException(
                        status_code=400,
                        detail="Legacy .doc files are not supported. Please convert to .docx format."
                    )
                
                doc = Document(io.BytesIO(content))
                extracted_text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
                
            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to read Word file: {str(e)}"
                )
        
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {file.content_type}"
            )
        
        # Parse the extracted text using the recipe parser
        parser = RecipeParser()
        parsed_recipe = parser.parse_recipe_text(extracted_text)
        result = parser.to_dict(parsed_recipe)
        
        return {
            "success": True,
            "extracted_text": extracted_text,
            "parsed_data": result,
            "file_type": file.content_type,
            "filename": file.filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process file: {str(e)}"
        ) 