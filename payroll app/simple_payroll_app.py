#!/usr/bin/env python3
"""
Simple Payroll App - Working Version
Bypasses the loading screen issue
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
import sqlite3
from datetime import datetime
import csv

# Fix for executable path issues
def get_app_directory():
    """Get the directory where the application is running from"""
    if getattr(sys, 'frozen', False):
        # Running as executable
        return os.path.dirname(sys.executable)
    else:
        # Running as script
        return os.path.dirname(os.path.abspath(__file__))

# Set up directories
APP_DIR = get_app_directory()
PROFILE_DIR = os.path.join(APP_DIR, 'profiles')

# Ensure profiles directory exists
if not os.path.exists(PROFILE_DIR):
    os.makedirs(PROFILE_DIR)

class SimplePayrollApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Professional Payroll System v2.0")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 800)
        
        # Configure modern styling
        self.setup_styling()
        
        # Initialize setup state
        self.setup_complete = False
        self.current_profile = None
        self.db_path = None
        
        # Check if setup is needed
        self.check_setup_status()
        
        if not self.setup_complete:
            self.show_setup_wizard()
        else:
            self.initialize_application()
    
    def setup_styling(self):
        """Configure modern styling for the application"""
        # Configure ttk styles
        style = ttk.Style()
        
        # Set theme
        try:
            style.theme_use('clam')  # Modern theme
        except:
            pass
        
        # Configure colors
        self.colors = {
            'primary': '#2E3B4E',
            'secondary': '#4CAF50', 
            'accent': '#2196F3',
            'background': '#F5F5F5',
            'surface': '#FFFFFF',
            'text': '#333333',
            'text_secondary': '#666666',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'error': '#F44336',
            'border': '#E0E0E0'
        }
        
        # Configure main window
        self.root.configure(bg=self.colors['background'])
        
        # Configure ttk styles
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 16, 'bold'),
                       foreground=self.colors['primary'])
        
        style.configure('Heading.TLabel',
                       font=('Segoe UI', 12, 'bold'),
                       foreground=self.colors['primary'])
        
        style.configure('Subheading.TLabel',
                       font=('Segoe UI', 10, 'bold'),
                       foreground=self.colors['text'])
        
        style.configure('Success.TLabel',
                       font=('Segoe UI', 10),
                       foreground=self.colors['success'])
        
        style.configure('Warning.TLabel',
                       font=('Segoe UI', 10),
                       foreground=self.colors['warning'])
        
        style.configure('Error.TLabel',
                       font=('Segoe UI', 10),
                       foreground=self.colors['error'])
        
        # Configure button styles
        style.configure('Primary.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white',
                       background=self.colors['accent'])
        
        style.configure('Success.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white',
                       background=self.colors['success'])
        
        style.configure('Warning.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white',
                       background=self.colors['warning'])
        
        style.configure('Error.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white',
                       background=self.colors['error'])
        
        # Configure frame styles
        style.configure('Card.TFrame',
                       background=self.colors['surface'],
                       relief='raised',
                       borderwidth=1)
        
        style.configure('Header.TFrame',
                       background=self.colors['primary'])
        
        # Configure notebook styles
        style.configure('TNotebook.Tab',
                       font=('Segoe UI', 10, 'bold'),
                       padding=[20, 10])
        
        style.configure('TNotebook',
                       background=self.colors['background'])
    
    def check_setup_status(self):
        """Check if application setup is complete"""
        try:
            # Check if there's a setup configuration file
            setup_file = os.path.join(PROFILE_DIR, 'setup_config.json')
            if os.path.exists(setup_file):
                import json
                with open(setup_file, 'r') as f:
                    config = json.load(f)
                    self.setup_complete = config.get('setup_complete', False)
                    self.current_profile = config.get('current_profile', 'Default')
                    self.db_path = os.path.join(PROFILE_DIR, f"{self.current_profile}.db")
            else:
                self.setup_complete = False
        except Exception as e:
            print(f"Error checking setup status: {e}")
            self.setup_complete = False
    
    def save_setup_config(self):
        """Save setup configuration"""
        try:
            import json
            setup_file = os.path.join(PROFILE_DIR, 'setup_config.json')
            config = {
                'setup_complete': True,
                'current_profile': self.current_profile,
                'setup_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            with open(setup_file, 'w') as f:
                json.dump(config, f, indent=2)
        except Exception as e:
            print(f"Error saving setup config: {e}")
    
    def show_setup_wizard(self):
        """Show the initial setup wizard"""
        # Create setup window
        setup_window = tk.Toplevel(self.root)
        setup_window.title("Business Setup - Professional Payroll System")
        setup_window.geometry("1000x800")
        setup_window.resizable(True, True)
        setup_window.transient(self.root)
        setup_window.grab_set()
        
        # Center the window
        setup_window.update_idletasks()
        x = (setup_window.winfo_screenwidth() // 2) - (1000 // 2)
        y = (setup_window.winfo_screenheight() // 2) - (800 // 2)
        setup_window.geometry(f"1000x800+{x}+{y}")
        
        # Main container
        main_container = tk.Frame(setup_window, bg=self.colors['background'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Header
        header_frame = tk.Frame(main_container, bg=self.colors['primary'], height=80)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="🏢 Welcome to Professional Payroll System",
                font=('Segoe UI', 18, 'bold'),
                fg='white', bg=self.colors['primary']).pack(pady=20)
        
        tk.Label(header_frame, text="Let's set up your business information",
                font=('Segoe UI', 12),
                fg='#B0C4DE', bg=self.colors['primary']).pack()
        
        # Progress indicator
        progress_frame = tk.Frame(header_frame, bg=self.colors['primary'])
        progress_frame.pack(pady=(5, 0))
        
        tk.Label(progress_frame, text="Step 1 of 1: Business Setup",
                font=('Segoe UI', 10),
                fg='#B0C4DE', bg=self.colors['primary']).pack()
        
        # Help text
        help_frame = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        help_frame.pack(fill=tk.X, pady=(0, 10))
        
        help_content = tk.Frame(help_frame, bg=self.colors['surface'])
        help_content.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(help_content, text="💡 Setup Guide:",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W)
        
        tk.Label(help_content, 
                text="• Fill in your business information (Business Name and Type are required)\n"
                     "• Configure your tax rates (defaults are provided)\n"
                     "• Click 'Complete Setup' to finish and start using your payroll system",
                font=('Segoe UI', 10),
                fg=self.colors['text'], bg=self.colors['surface'],
                justify=tk.LEFT).pack(anchor=tk.W, pady=(5, 0))
        
        # Setup form with scrollable content
        form_card = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        form_card.pack(fill=tk.BOTH, expand=True)
        
        # Create canvas and scrollbar for scrollable content
        canvas = tk.Canvas(form_card, bg=self.colors['surface'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(form_card, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['surface'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            try:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            except:
                pass  # Ignore errors if canvas is destroyed
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Clean up mousewheel binding when window closes
        def _unbind_mousewheel():
            try:
                canvas.unbind_all("<MouseWheel>")
            except:
                pass  # Ignore errors during cleanup
        setup_window.protocol("WM_DELETE_WINDOW", _unbind_mousewheel)
        
        form_content = scrollable_frame
        
        # Add padding to form content
        form_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Business Information Section
        business_frame = tk.Frame(form_content, bg=self.colors['surface'])
        business_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(business_frame, text="📋 Business Information",
                font=('Segoe UI', 14, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W, pady=(0, 15))
        
        # Business fields (reduced for better fit)
        fields = [
            ("Business Name", "business_name_entry", True, "Enter your business name"),
            ("Business Type", "business_type_entry", True, "e.g., Restaurant, Retail, Service"),
            ("Address", "business_address_entry", False, "Street address"),
            ("City, State, ZIP", "city_entry", False, "City, State, ZIP"),
            ("Phone", "phone_entry", False, "Business phone"),
            ("Email", "email_entry", False, "Business email")
        ]
        
        self.setup_entries = {}
        
        for label, entry_name, required, placeholder in fields:
            field_frame = tk.Frame(business_frame, bg=self.colors['surface'])
            field_frame.pack(fill=tk.X, pady=8)
            
            # Label
            label_text = f"{label}:" + (" *" if required else "")
            label_color = self.colors['error'] if required else self.colors['text']
            
            tk.Label(field_frame, text=label_text,
                    font=('Segoe UI', 10, 'bold'),
                    fg=label_color, bg=self.colors['surface']).pack(anchor=tk.W)
            
            # Entry
            entry = tk.Entry(field_frame, font=('Segoe UI', 10), width=50, 
                           fg=self.colors['text_secondary'], relief='solid', bd=1)
            entry.pack(fill=tk.X, pady=(5, 0))
            entry.insert(0, placeholder)
            entry.bind('<FocusIn>', lambda e, entry=entry, placeholder=placeholder: self.clear_placeholder(entry, placeholder))
            entry.bind('<FocusOut>', lambda e, entry=entry, placeholder=placeholder: self.restore_placeholder(entry, placeholder))
            
            self.setup_entries[entry_name] = entry
        
        # Tax Configuration Section
        tax_frame = tk.Frame(form_content, bg=self.colors['surface'])
        tax_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(tax_frame, text="💰 Tax Configuration",
                font=('Segoe UI', 14, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W, pady=(0, 15))
        
        # Tax rate fields
        tax_fields_frame = tk.Frame(tax_frame, bg=self.colors['surface'])
        tax_fields_frame.pack(fill=tk.X)
        
        # Federal Tax
        federal_frame = tk.Frame(tax_fields_frame, bg=self.colors['surface'])
        federal_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(federal_frame, text="Federal Tax Rate (%):",
                font=('Segoe UI', 10, 'bold'),
                fg=self.colors['text'], bg=self.colors['surface']).pack(side=tk.LEFT)
        
        self.federal_rate_entry = tk.Entry(federal_frame, font=('Segoe UI', 10), width=10,
                                          fg=self.colors['text'], relief='solid', bd=1)
        self.federal_rate_entry.pack(side=tk.RIGHT)
        self.federal_rate_entry.insert(0, "15.0")
        
        # State Tax
        state_frame = tk.Frame(tax_fields_frame, bg=self.colors['surface'])
        state_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(state_frame, text="State Tax Rate (%):",
                font=('Segoe UI', 10, 'bold'),
                fg=self.colors['text'], bg=self.colors['surface']).pack(side=tk.LEFT)
        
        self.state_rate_entry = tk.Entry(state_frame, font=('Segoe UI', 10), width=10,
                                        fg=self.colors['text'], relief='solid', bd=1)
        self.state_rate_entry.pack(side=tk.RIGHT)
        self.state_rate_entry.insert(0, "5.0")
        
        # Buttons section (outside scrollable area, always visible)
        button_frame = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        button_content = tk.Frame(button_frame, bg=self.colors['surface'])
        button_content.pack(fill=tk.X, padx=30, pady=20)
        
        # Cancel button
        cancel_btn = tk.Button(button_content, text="← Cancel", 
                              command=setup_window.destroy,
                              font=('Segoe UI', 10), 
                              bg=self.colors['text_secondary'], fg='white',
                              relief='flat', padx=20, pady=10, cursor='hand2')
        cancel_btn.pack(side=tk.LEFT)
        
        # Complete Setup button (more prominent)
        complete_btn = tk.Button(button_content, text="✅ Complete Setup", 
                                command=lambda: self.complete_setup(setup_window),
                                font=('Segoe UI', 12, 'bold'), 
                                bg=self.colors['success'], fg='white',
                                relief='flat', padx=30, pady=12, cursor='hand2')
        complete_btn.pack(side=tk.RIGHT)
        
        # Add some spacing
        tk.Frame(button_content, bg=self.colors['surface'], width=20).pack(side=tk.RIGHT)
    
    def clear_placeholder(self, entry, placeholder):
        """Clear placeholder text when entry is focused"""
        if entry.get() == placeholder:
            entry.delete(0, tk.END)
            entry.config(fg=self.colors['text'])
    
    def restore_placeholder(self, entry, placeholder):
        """Restore placeholder text if entry is empty"""
        if not entry.get():
            entry.insert(0, placeholder)
            entry.config(fg=self.colors['text_secondary'])
    
    def complete_setup(self, setup_window):
        """Complete the setup process"""
        try:
            # Validate required fields
            business_name = self.setup_entries['business_name_entry'].get()
            business_type = self.setup_entries['business_type_entry'].get()
            
            if business_name == "Enter your business name" or not business_name.strip():
                messagebox.showerror("Setup Error", "Please enter your business name")
                return
            
            if business_type == "e.g., Restaurant, Retail, Service" or not business_type.strip():
                messagebox.showerror("Setup Error", "Please enter your business type")
                return
            
            # Set profile name based on business name
            profile_name = business_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
            self.current_profile = profile_name
            self.db_path = os.path.join(PROFILE_DIR, f"{profile_name}.db")
            
            # Save setup configuration
            self.save_setup_config()
            
            # Setup performance features first
            self.setup_performance_features()
            
            # Setup database
            self.setup_database()
            
            # Save initial settings
            self.save_initial_settings()
            
            # Close setup window
            setup_window.destroy()
            
            # Initialize the main application
            self.initialize_application()
            
            messagebox.showinfo("Setup Complete", 
                f"Welcome to Professional Payroll System!\n\nBusiness: {business_name}\nProfile: {profile_name}\n\nYour payroll system is ready to use!")
            
        except Exception as e:
            messagebox.showerror("Setup Error", f"Failed to complete setup: {e}")
    
    def save_initial_settings(self):
        """Save initial settings from setup"""
        try:
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            
            # Create settings table if it doesn't exist
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            ''')
            
            # Save business information
            business_info = {
                'business_name': self.setup_entries['business_name_entry'].get(),
                'business_type': self.setup_entries['business_type_entry'].get(),
                'business_address': self.setup_entries['business_address_entry'].get(),
                'city': self.setup_entries['city_entry'].get(),
                'state': '',  # Combined with city
                'zip_code': '',  # Combined with city
                'phone': self.setup_entries['phone_entry'].get(),
                'email': self.setup_entries['email_entry'].get(),
                'federal_tax_rate': self.federal_rate_entry.get(),
                'state_tax_rate': self.state_rate_entry.get(),
                'ss_tax_rate': '6.2',
                'medicare_tax_rate': '1.45',
                'pfml_tax_rate': '0.75',
                'meals_tax_rate': '0.0',
                'service_tax_rate': '0.0',
                'pay_frequency': 'Weekly'
            }
            
            for key, value in business_info.items():
                cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                              (key, value))
            
            conn.commit()
            
        except Exception as e:
            print(f"Error saving initial settings: {e}")
    
    def initialize_application(self):
        """Initialize the main application after setup"""
        self.setup_database()
        self.create_interface()
    
    def setup_database(self):
        """Set up the SQLite database with performance optimizations"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Performance optimizations
            cursor.execute("PRAGMA journal_mode = WAL")  # Better concurrency
            cursor.execute("PRAGMA synchronous = NORMAL")  # Faster writes
            cursor.execute("PRAGMA cache_size = 10000")  # Larger cache
            cursor.execute("PRAGMA temp_store = MEMORY")  # Store temp in memory
            cursor.execute("PRAGMA mmap_size = 268435456")  # 256MB memory mapping
            
            # Create employees table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS employees (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    position TEXT,
                    hourly_rate REAL NOT NULL,
                    hire_date TEXT,
                    phone TEXT,
                    email TEXT
                )
            ''')
            
            # Create payroll_runs table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS payroll_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    run_date TEXT NOT NULL,
                    frequency TEXT NOT NULL,
                    hours_worked REAL NOT NULL,
                    gross_pay REAL NOT NULL,
                    federal_tax REAL DEFAULT 0.0,
                    ss_tax REAL DEFAULT 0.0,
                    medicare_tax REAL DEFAULT 0.0,
                    state_tax REAL DEFAULT 0.0,
                    meals_tax REAL DEFAULT 0.0,
                    service_tax REAL DEFAULT 0.0,
                    pfml_tax REAL DEFAULT 0.0,
                    total_taxes REAL NOT NULL,
                    net_pay REAL NOT NULL,
                    bonus REAL DEFAULT 0.0,
                    FOREIGN KEY (employee_id) REFERENCES employees(id)
                )
            ''')
            
            # Create indexes for better performance
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_employees_name ON employees(name)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_payroll_employee ON payroll_runs(employee_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_payroll_date ON payroll_runs(run_date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_settings_key ON settings(key)")
            
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to setup database: {e}")
    
    def create_interface(self):
        """Create the main interface with modern design"""
        # Create main container
        main_container = tk.Frame(self.root, bg=self.colors['background'])
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Create header
        self.create_header(main_container)
        
        # Create content area
        content_frame = tk.Frame(main_container, bg=self.colors['background'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        self.create_dashboard_tab()
        self.create_employees_tab()
        self.create_payroll_tab()
        self.create_reports_tab()
        self.create_settings_tab()
        self.create_help_tab()
        
        # Performance optimizations (must be first)
        self.setup_performance_features()
        
        # Load initial data
        self.load_employees()
        
        # Setup keyboard shortcuts
        self.setup_keyboard_shortcuts()
    
    def setup_performance_features(self):
        """Setup performance optimization features"""
        # Connection caching
        self._db_connection = None
        self._last_db_access = 0
        self._connection_timeout = 30  # 30 seconds
        
        # Data caching
        self._employee_cache = {}
        self._settings_cache = {}
        self._cache_timeout = 300  # 5 minutes
    
    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts for common actions"""
        # Ctrl+S to save (when in appropriate context)
        self.root.bind('<Control-s>', self.quick_save)
        
        # Ctrl+N to add new employee
        self.root.bind('<Control-n>', self.quick_new_employee)
        
        # Ctrl+1-5 to switch tabs
        self.root.bind('<Control-1>', lambda e: self.notebook.select(0))  # Dashboard
        self.root.bind('<Control-2>', lambda e: self.notebook.select(1))  # Employees
        self.root.bind('<Control-3>', lambda e: self.notebook.select(2))  # Payroll
        self.root.bind('<Control-4>', lambda e: self.notebook.select(3))  # Reports
        self.root.bind('<Control-5>', lambda e: self.notebook.select(4))  # Settings
        
        # F5 to refresh data
        self.root.bind('<F5>', self.quick_refresh)
        
        # Escape to clear search
        self.root.bind('<Escape>', self.quick_clear_search)
    
    def quick_save(self, event):
        """Quick save action"""
        try:
            current_tab = self.notebook.index(self.notebook.select())
            if current_tab == 2:  # Payroll tab
                self.save_payroll()
            elif current_tab == 4:  # Settings tab
                self.save_settings()
            else:
                self.update_status("No save action available on this tab", "info")
        except Exception as e:
            print(f"Error in quick save: {e}")
    
    def quick_new_employee(self, event):
        """Quick new employee action"""
        self.notebook.select(1)  # Switch to employees tab
        self.update_status("Ready to add new employee (Ctrl+N)", "info")
    
    def quick_refresh(self, event):
        """Quick refresh action"""
        self.load_employees()
        self.refresh_dashboard()
        self.update_status("Data refreshed (F5)", "success")
    
    def quick_clear_search(self, event):
        """Quick clear search action"""
        try:
            if hasattr(self, 'employee_search_var'):
                self.employee_search_var.set("")
                self.update_status("Search cleared (Escape)", "info")
        except Exception as e:
            print(f"Error clearing search: {e}")
    
    def get_optimized_connection(self):
        """Get optimized database connection with caching"""
        import time
        
        current_time = time.time()
        
        # Reuse connection if recent
        if (self._db_connection is not None and 
            current_time - self._last_db_access < self._connection_timeout):
            self._last_db_access = current_time
            return self._db_connection
        
        # Create new connection with optimizations
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Apply performance settings
        cursor.execute("PRAGMA journal_mode = WAL")
        cursor.execute("PRAGMA synchronous = NORMAL")
        cursor.execute("PRAGMA cache_size = 10000")
        cursor.execute("PRAGMA temp_store = MEMORY")
        
        self._db_connection = conn
        self._last_db_access = current_time
        
        return conn
    
    def get_business_name(self):
        """Get business name from settings"""
        try:
            if hasattr(self, '_settings_cache') and 'business_name' in self._settings_cache:
                return self._settings_cache['business_name']
            
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key = 'business_name'")
            result = cursor.fetchone()
            if result:
                business_name = result[0]
                self._settings_cache['business_name'] = business_name
                return business_name
            return None
        except Exception as e:
            print(f"Error getting business name: {e}")
            return None
    
    def create_header(self, parent):
        """Create modern header with branding and status"""
        header_frame = tk.Frame(parent, bg=self.colors['primary'], height=80)
        header_frame.pack(fill=tk.X, padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        # Left side - Logo and title
        left_frame = tk.Frame(header_frame, bg=self.colors['primary'])
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=10)
        
        # Title
        business_name = self.get_business_name()
        title_text = f"🏢 {business_name}" if business_name else "🏢 Professional Payroll System"
        title_label = tk.Label(left_frame, 
                              text=title_text,
                              font=('Segoe UI', 20, 'bold'),
                              fg='white',
                              bg=self.colors['primary'])
        title_label.pack(anchor=tk.W)
        
        # Subtitle
        subtitle_text = "Professional Payroll Management System" if business_name else "Complete Payroll Management Solution"
        subtitle_label = tk.Label(left_frame,
                                 text=subtitle_text,
                                 font=('Segoe UI', 11),
                                 fg='#B0C4DE',
                                 bg=self.colors['primary'])
        subtitle_label.pack(anchor=tk.W)
        
        # Right side - Status and info
        right_frame = tk.Frame(header_frame, bg=self.colors['primary'])
        right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=20, pady=10)
        
        # Version info
        # Save status
        self.save_status_label = tk.Label(right_frame,
                                        text="💾 Saved",
                                        font=('Segoe UI', 9),
                                        fg='#B0C4DE',
                                        bg=self.colors['primary'])
        self.save_status_label.pack(anchor=tk.E)
        
        # Last saved time
        self.last_saved_label = tk.Label(right_frame,
                                        text="",
                                        font=('Segoe UI', 8),
                                        fg='#B0C4DE',
                                        bg=self.colors['primary'])
        self.last_saved_label.pack(anchor=tk.E)
        
        # Status
        self.status_label = tk.Label(right_frame,
                                    text="✅ Ready",
                                    font=('Segoe UI', 10, 'bold'),
                                    fg=self.colors['success'],
                                    bg=self.colors['primary'])
        self.status_label.pack(anchor=tk.E, pady=(5, 0))
        
        # Version info
        version_label = tk.Label(right_frame,
                                text="Version 2.0.0",
                                font=('Segoe UI', 9),
                                fg='#B0C4DE',
                                bg=self.colors['primary'])
        version_label.pack(anchor=tk.E, pady=(5, 0))
    
    def update_status(self, message, status_type="success"):
        """Update the status message in the header"""
        colors = {
            'success': self.colors['success'],
            'warning': self.colors['warning'],
            'error': self.colors['error'],
            'info': self.colors['accent']
        }
        
        icons = {
            'success': '✅',
            'warning': '⚠️',
            'error': '❌',
            'info': 'ℹ️'
        }
        
        self.status_label.config(
            text=f"{icons.get(status_type, 'ℹ️')} {message}",
            fg=colors.get(status_type, self.colors['text'])
        )
    
    def update_save_status(self, status="saved"):
        """Update save status indicator"""
        from datetime import datetime
        
        if status == "saved":
            self.save_status_label.config(text="💾 Saved", fg=self.colors['success'])
            self.last_saved_label.config(text=f"Last: {datetime.now().strftime('%H:%M:%S')}")
        elif status == "unsaved":
            self.save_status_label.config(text="⚠️ Unsaved", fg=self.colors['warning'])
            self.last_saved_label.config(text="")
        elif status == "saving":
            self.save_status_label.config(text="💾 Saving...", fg=self.colors['accent'])
            self.last_saved_label.config(text="")
    
    def show_loading_indicator(self, message="Loading..."):
        """Show a simple loading indicator"""
        self.update_status(f"⏳ {message}", "info")
    
    def show_success_animation(self, message="Success!"):
        """Show success animation"""
        self.update_status(f"✅ {message}", "success")
        # Auto-clear after 3 seconds
        self.root.after(3000, lambda: self.update_status("Ready", "success"))
    
    def create_dashboard_tab(self):
        """Create the dashboard tab with key metrics and quick actions"""
        dashboard_frame = ttk.Frame(self.notebook)
        self.notebook.add(dashboard_frame, text="📊 Dashboard")
        
        # Main container with padding
        main_container = tk.Frame(dashboard_frame, bg=self.colors['background'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header
        header_frame = tk.Frame(main_container, bg=self.colors['primary'], relief='raised', bd=1)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        header_content = tk.Frame(header_frame, bg=self.colors['primary'])
        header_content.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Label(header_content, text="📊 Business Overview",
                font=('Segoe UI', 16, 'bold'),
                fg='white', bg=self.colors['primary']).pack(anchor=tk.W)
        
        tk.Label(header_content, text="Key metrics and quick actions for your payroll system",
                font=('Segoe UI', 11),
                fg='#B0C4DE', bg=self.colors['primary']).pack(anchor=tk.W, pady=(5, 0))
        
        # Metrics cards container
        metrics_container = tk.Frame(main_container, bg=self.colors['background'])
        metrics_container.pack(fill=tk.X, pady=(0, 20))
        
        # Employee count card
        emp_card = tk.Frame(metrics_container, bg=self.colors['surface'], relief='raised', bd=1)
        emp_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        emp_content = tk.Frame(emp_card, bg=self.colors['surface'])
        emp_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        tk.Label(emp_content, text="👥 Total Employees",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W)
        
        self.employee_count_label = tk.Label(emp_content, text="0",
                font=('Segoe UI', 24, 'bold'),
                fg=self.colors['success'], bg=self.colors['surface'])
        self.employee_count_label.pack(anchor=tk.W, pady=(10, 0))
        
        # Recent payroll card
        payroll_card = tk.Frame(metrics_container, bg=self.colors['surface'], relief='raised', bd=1)
        payroll_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        payroll_content = tk.Frame(payroll_card, bg=self.colors['surface'])
        payroll_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        tk.Label(payroll_content, text="💰 Recent Payroll",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W)
        
        self.recent_payroll_label = tk.Label(payroll_content, text="No payroll yet",
                font=('Segoe UI', 14),
                fg=self.colors['text'], bg=self.colors['surface'])
        self.recent_payroll_label.pack(anchor=tk.W, pady=(10, 0))
        
        # Monthly total card
        monthly_card = tk.Frame(metrics_container, bg=self.colors['surface'], relief='raised', bd=1)
        monthly_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        monthly_content = tk.Frame(monthly_card, bg=self.colors['surface'])
        monthly_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        tk.Label(monthly_content, text="📅 This Month",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W)
        
        self.monthly_total_label = tk.Label(monthly_content, text="$0.00",
                font=('Segoe UI', 20, 'bold'),
                fg=self.colors['success'], bg=self.colors['surface'])
        self.monthly_total_label.pack(anchor=tk.W, pady=(10, 0))
        
        # Quick actions section
        actions_frame = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        actions_frame.pack(fill=tk.X, pady=(0, 20))
        
        actions_content = tk.Frame(actions_frame, bg=self.colors['surface'])
        actions_content.pack(fill=tk.X, padx=20, pady=20)
        
        tk.Label(actions_content, text="⚡ Quick Actions",
                font=('Segoe UI', 14, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W, pady=(0, 15))
        
        # Quick action buttons
        actions_row = tk.Frame(actions_content, bg=self.colors['surface'])
        actions_row.pack(fill=tk.X)
        
        tk.Button(actions_row, text="➕ Add Employee",
                command=self.quick_add_employee,
                font=('Segoe UI', 11, 'bold'), bg=self.colors['success'], fg='white',
                relief='flat', padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(actions_row, text="💰 Process Payroll",
                command=self.quick_process_payroll,
                font=('Segoe UI', 11, 'bold'), bg=self.colors['accent'], fg='white',
                relief='flat', padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(actions_row, text="📊 View Reports",
                command=self.quick_view_reports,
                font=('Segoe UI', 11, 'bold'), bg=self.colors['secondary'], fg='white',
                relief='flat', padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(actions_row, text="⚙️ Settings",
                command=self.quick_open_settings,
                font=('Segoe UI', 11, 'bold'), bg=self.colors['text_secondary'], fg='white',
                relief='flat', padx=20, pady=10, cursor='hand2').pack(side=tk.LEFT)
        
        # Recent activity section
        activity_frame = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        activity_frame.pack(fill=tk.BOTH, expand=True)
        
        activity_content = tk.Frame(activity_frame, bg=self.colors['surface'])
        activity_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        tk.Label(activity_content, text="📋 Recent Activity",
                font=('Segoe UI', 14, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(anchor=tk.W, pady=(0, 15))
        
        # Activity list
        self.activity_listbox = tk.Listbox(activity_content, height=8, font=('Segoe UI', 10),
                                         bg=self.colors['surface'], fg=self.colors['text'],
                                         selectbackground=self.colors['accent'], relief='flat')
        self.activity_listbox.pack(fill=tk.BOTH, expand=True)
        
        # Load dashboard data
        self.refresh_dashboard()
    
    def refresh_dashboard(self):
        """Refresh dashboard data"""
        try:
            # Get employee count
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM employees")
            emp_count = cursor.fetchone()[0]
            self.employee_count_label.config(text=str(emp_count))
            
            # Get recent payroll
            cursor.execute("""
                SELECT pr.run_date, e.name, pr.net_pay
                FROM payroll_runs pr
                JOIN employees e ON pr.employee_id = e.id
                ORDER BY pr.run_date DESC
                LIMIT 1
            """)
            recent = cursor.fetchone()
            if recent:
                self.recent_payroll_label.config(text=f"{recent[1]} - ${recent[2]:.2f} ({recent[0]})")
            else:
                self.recent_payroll_label.config(text="No payroll yet")
            
            # Get monthly total
            current_month = datetime.now().strftime('%Y-%m')
            cursor.execute("""
                SELECT SUM(net_pay)
                FROM payroll_runs
                WHERE run_date LIKE ?
            """, (f"{current_month}%",))
            monthly_total = cursor.fetchone()[0] or 0
            self.monthly_total_label.config(text=f"${monthly_total:.2f}")
            
            # Update activity list
            self.update_activity_list()
            
        except Exception as e:
            print(f"Error refreshing dashboard: {e}")
    
    def update_activity_list(self):
        """Update the recent activity list"""
        try:
            self.activity_listbox.delete(0, tk.END)
            
            # Get recent payroll runs
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT pr.run_date, e.name, pr.net_pay
                FROM payroll_runs pr
                JOIN employees e ON pr.employee_id = e.id
                ORDER BY pr.run_date DESC
                LIMIT 10
            """)
            payroll_runs = cursor.fetchall()
            
            for run in payroll_runs:
                self.activity_listbox.insert(tk.END, f"💰 Payroll: {run[1]} - ${run[2]:.2f} ({run[0]})")
            
            # Add employee activities if any
            cursor.execute("""
                SELECT name, 'Added' as action, 'N/A' as date
                FROM employees
                ORDER BY id DESC
                LIMIT 5
            """)
            employees = cursor.fetchall()
            
            for emp in employees:
                self.activity_listbox.insert(tk.END, f"👤 Employee: {emp[0]} {emp[1]}")
                
        except Exception as e:
            print(f"Error updating activity list: {e}")
    
    def quick_add_employee(self):
        """Quick action to add employee"""
        self.notebook.select(1)  # Switch to employees tab
        self.update_status("Switched to Employees tab - ready to add new employee", "info")
    
    def quick_process_payroll(self):
        """Quick action to process payroll"""
        self.notebook.select(2)  # Switch to payroll tab
        self.update_status("Switched to Payroll tab - ready to process payroll", "info")
    
    def quick_view_reports(self):
        """Quick action to view reports"""
        self.notebook.select(3)  # Switch to reports tab
        self.update_status("Switched to Reports tab", "info")
    
    def quick_open_settings(self):
        """Quick action to open settings"""
        self.notebook.select(4)  # Switch to settings tab
        self.update_status("Switched to Settings tab", "info")
    
    def create_employees_tab(self):
        """Create the employees management tab with modern design"""
        employees_frame = ttk.Frame(self.notebook)
        self.notebook.add(employees_frame, text="👥 Employees")
        
        # Create main container with padding
        main_container = tk.Frame(employees_frame, bg=self.colors['background'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Left side - Employee list with card design
        list_card = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        list_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # List header
        list_header = tk.Frame(list_card, bg=self.colors['primary'], height=40)
        list_header.pack(fill=tk.X)
        list_header.pack_propagate(False)
        
        tk.Label(list_header, text="📋 Employee Directory", 
                font=('Segoe UI', 12, 'bold'),
                fg='white', bg=self.colors['primary']).pack(pady=10)
        
        # List content
        list_content = tk.Frame(list_card, bg=self.colors['surface'])
        list_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Search section
        search_frame = tk.Frame(list_content, bg=self.colors['surface'])
        search_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(search_frame, text="🔍 Search:",
                font=('Segoe UI', 10, 'bold'),
                fg=self.colors['primary'], bg=self.colors['surface']).pack(side=tk.LEFT)
        
        self.employee_search_var = tk.StringVar()
        self.employee_search_var.trace('w', self.filter_employees)
        search_entry = tk.Entry(search_frame, textvariable=self.employee_search_var,
                              font=('Segoe UI', 10), width=25, relief='solid', bd=1)
        search_entry.pack(side=tk.LEFT, padx=(10, 5))
        
        # Clear search button
        tk.Button(search_frame, text="✖️",
                command=self.clear_employee_search,
                font=('Segoe UI', 9), bg=self.colors['text_secondary'], fg='white',
                relief='flat', padx=8, pady=2, cursor='hand2').pack(side=tk.LEFT)
        
        # Employee listbox with modern styling
        self.employee_listbox = tk.Listbox(list_content, height=15, 
                                          font=('Segoe UI', 10),
                                          bg=self.colors['surface'],
                                          fg=self.colors['text'],
                                          selectbackground=self.colors['accent'],
                                          selectforeground='white',
                                          relief='flat',
                                          bd=1)
        self.employee_listbox.pack(fill=tk.BOTH, expand=True)
        self.employee_listbox.bind('<<ListboxSelect>>', self.on_employee_select)
    
    def filter_employees(self, *args):
        """Filter employees based on search text"""
        try:
            search_text = self.employee_search_var.get().lower()
            
            # Clear current list
            self.employee_listbox.delete(0, tk.END)
            
            # Get all employees from cache or database
            if hasattr(self, '_employee_cache') and self._employee_cache:
                employees = self._employee_cache
            else:
                conn = self.get_optimized_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT name, position, hourly_rate, phone, email FROM employees ORDER BY name")
                employees = cursor.fetchall()
            
            # Filter and display employees
            filtered_count = 0
            for emp in employees:
                name, position, rate, phone, email = emp
                display_text = f"{name} - {position} (${rate:.2f}/hr)"
                
                if search_text in name.lower() or search_text in position.lower() or search_text in phone.lower() or search_text in email.lower():
                    self.employee_listbox.insert(tk.END, display_text)
                    filtered_count += 1
            
            # Update status
            if search_text:
                self.update_status(f"Found {filtered_count} employees matching '{search_text}'", "info")
            else:
                self.update_status(f"Showing {filtered_count} employees", "info")
                
        except Exception as e:
            print(f"Error filtering employees: {e}")
    
    def clear_employee_search(self):
        """Clear employee search"""
        self.employee_search_var.set("")
        self.load_employees()
        self.update_status("Employee search cleared", "info")
    
    def load_employees_lazy(self):
        """Load employees with lazy loading for better performance"""
        try:
            # Show loading indicator
            self.update_status("Loading employees...", "info")
            
            # Use optimized connection
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            
            # Check cache first
            import time
            current_time = time.time()
            if 'employees' in self._employee_cache:
                cache_time, employees = self._employee_cache['employees']
                if current_time - cache_time < self._cache_timeout:
                    self.populate_employee_list(employees)
                    self.update_status("Employees loaded from cache")
                    return
            
            # Load from database
            cursor.execute("SELECT id, name, position, hourly_rate FROM employees ORDER BY name")
            employees = cursor.fetchall()
            
            # Cache the results
            self._employee_cache['employees'] = (current_time, employees)
            
            # Populate the list
            self.populate_employee_list(employees)
            self.update_status(f"Loaded {len(employees)} employees")
            
        except Exception as e:
            self.update_status(f"Error loading employees: {e}", "error")
            messagebox.showerror("Error", f"Failed to load employees: {e}")
    
    def populate_employee_list(self, employees):
        """Populate employee list efficiently"""
        self.employee_listbox.delete(0, tk.END)
        
        for employee in employees:
            emp_id, name, position, rate = employee
            display_text = f"{name} - {position} (${rate:.2f}/hr)"
            self.employee_listbox.insert(tk.END, display_text)
        
        # Right side - Employee form with card design
        form_card = tk.Frame(main_container, bg=self.colors['surface'], relief='raised', bd=1)
        form_card.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        
        # Form header
        form_header = tk.Frame(form_card, bg=self.colors['primary'], height=40)
        form_header.pack(fill=tk.X)
        form_header.pack_propagate(False)
        
        tk.Label(form_header, text="✏️ Employee Details", 
                font=('Segoe UI', 12, 'bold'),
                fg='white', bg=self.colors['primary']).pack(pady=10)
        
        # Form content
        form_content = tk.Frame(form_card, bg=self.colors['surface'])
        form_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Form fields with modern styling
        fields = [
            ("Name", "name_entry", True),
            ("Position", "position_entry", False),
            ("Hourly Rate", "rate_entry", True),
            ("Phone", "phone_entry", False),
            ("Email", "email_entry", False)
        ]
        
        for i, (label, entry_name, required) in enumerate(fields):
            field_frame = tk.Frame(form_content, bg=self.colors['surface'])
            field_frame.pack(fill=tk.X, pady=8)
            
            # Label with required indicator
            label_text = f"{label}:" + (" *" if required else "")
            label_color = self.colors['error'] if required else self.colors['text']
            
            tk.Label(field_frame, text=label_text,
                    font=('Segoe UI', 10, 'bold'),
                    fg=label_color, bg=self.colors['surface']).pack(anchor=tk.W)
            
            # Entry field
            entry = ttk.Entry(field_frame, width=25, font=('Segoe UI', 10))
            entry.pack(fill=tk.X, pady=(5, 0))
            setattr(self, entry_name, entry)
        
        # Action buttons with modern styling
        button_frame = tk.Frame(form_content, bg=self.colors['surface'])
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(button_frame, text="➕ Add Employee", 
                  command=self.add_employee, style='Success.TButton').pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="✏️ Update Employee", 
                  command=self.update_employee, style='Primary.TButton').pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="🗑️ Delete Employee", 
                  command=self.delete_employee, style='Error.TButton').pack(fill=tk.X, pady=5)
        ttk.Button(button_frame, text="🧹 Clear Form", 
                  command=self.clear_form, style='Warning.TButton').pack(fill=tk.X, pady=5)
    
    def create_payroll_tab(self):
        """Create the payroll processing tab with modern design"""
        payroll_frame = ttk.Frame(self.notebook)
        self.notebook.add(payroll_frame, text="💰 Payroll")
        
        # Left side - Payroll input
        input_frame = ttk.LabelFrame(payroll_frame, text="Payroll Input", padding="10")
        input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        ttk.Label(input_frame, text="Employee:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.payroll_employee_var = tk.StringVar()
        self.payroll_employee_combo = ttk.Combobox(input_frame, textvariable=self.payroll_employee_var, width=25)
        self.payroll_employee_combo.grid(row=0, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(input_frame, text="Hours Worked:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.hours_entry = ttk.Entry(input_frame, width=25)
        self.hours_entry.grid(row=1, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(input_frame, text="Hourly Rate:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.payroll_rate_entry = ttk.Entry(input_frame, width=25)
        self.payroll_rate_entry.grid(row=2, column=1, sticky=tk.W, pady=2)
        
        # Payroll Date
        ttk.Label(input_frame, text="Payroll Date:").grid(row=3, column=0, sticky=tk.W, pady=2)
        date_frame = tk.Frame(input_frame)
        date_frame.grid(row=3, column=1, sticky=tk.W, pady=2)
        
        self.payroll_date_entry = tk.Entry(date_frame, font=('Segoe UI', 10), width=15, relief='solid', bd=1)
        self.payroll_date_entry.pack(side=tk.LEFT)
        self.payroll_date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Quick date buttons
        quick_dates_frame = tk.Frame(input_frame)
        quick_dates_frame.grid(row=4, column=1, sticky=tk.W, pady=5)
        
        tk.Button(quick_dates_frame, text="Today", command=lambda: self.set_quick_date("today"),
                font=('Segoe UI', 8), bg=self.colors['accent'], fg='white',
                relief='flat', padx=8, pady=2, cursor='hand2').pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Button(quick_dates_frame, text="Last Week", command=lambda: self.set_quick_date("last_week"),
                font=('Segoe UI', 8), bg=self.colors['secondary'], fg='white',
                relief='flat', padx=8, pady=2, cursor='hand2').pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Button(quick_dates_frame, text="This Month", command=lambda: self.set_quick_date("this_month"),
                font=('Segoe UI', 8), bg=self.colors['success'], fg='white',
                relief='flat', padx=8, pady=2, cursor='hand2').pack(side=tk.LEFT)
        
        # Calculate button
        ttk.Button(input_frame, text="Calculate Payroll", command=self.calculate_payroll).grid(row=5, column=0, columnspan=2, pady=20)
        
        # Right side - Results
        results_frame = ttk.LabelFrame(payroll_frame, text="Payroll Results", padding="10")
        results_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        ttk.Label(results_frame, text="Gross Pay:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.gross_pay_label = ttk.Label(results_frame, text="$0.00")
        self.gross_pay_label.grid(row=0, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(results_frame, text="Taxes:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.taxes_label = ttk.Label(results_frame, text="$0.00")
        self.taxes_label.grid(row=1, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(results_frame, text="Net Pay:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.net_pay_label = ttk.Label(results_frame, text="$0.00")
        self.net_pay_label.grid(row=2, column=1, sticky=tk.W, pady=2)
        
        # Save button
        ttk.Button(results_frame, text="Save Payroll", command=self.save_payroll).grid(row=3, column=0, columnspan=2, pady=20)
        
        # Payroll History Section
        history_frame = ttk.LabelFrame(payroll_frame, text="Recent Payroll Runs")
        history_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Create treeview for payroll history
        columns = ("ID", "Employee", "Date", "Hours", "Gross Pay", "Net Pay")
        self.payroll_tree = ttk.Treeview(history_frame, columns=columns, show="headings", height=6)
        
        # Configure columns
        for col in columns:
            self.payroll_tree.heading(col, text=col)
            self.payroll_tree.column(col, width=100)
        
        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.payroll_tree.yview)
        self.payroll_tree.configure(yscrollcommand=scrollbar.set)
        
        self.payroll_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons for payroll management
        button_frame = ttk.Frame(history_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(button_frame, text="Refresh", command=self.load_payroll_history).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Edit Selected", command=self.edit_selected_payroll).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete Selected", command=self.delete_selected_payroll).pack(side=tk.LEFT, padx=5)
        
        # Load initial payroll history
        self.load_payroll_history()
    
    def set_quick_date(self, date_type):
        """Set quick date for payroll"""
        from datetime import datetime, timedelta
        
        if date_type == "today":
            date_str = datetime.now().strftime("%Y-%m-%d")
        elif date_type == "last_week":
            date_str = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        elif date_type == "this_month":
            date_str = datetime.now().strftime("%Y-%m-01")
        else:
            date_str = datetime.now().strftime("%Y-%m-%d")
        
        self.payroll_date_entry.delete(0, tk.END)
        self.payroll_date_entry.insert(0, date_str)
        self.update_status(f"Date set to {date_str}", "info")
    
    def create_reports_tab(self):
        """Create the reports tab"""
        reports_frame = ttk.Frame(self.notebook)
        self.notebook.add(reports_frame, text="📊 Reports")
        
        # Title
        ttk.Label(reports_frame, text="Payroll Reports & Export", font=("Arial", 16, "bold")).pack(pady=20)
        
        # Create notebook for different report types
        report_notebook = ttk.Notebook(reports_frame)
        report_notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Employee Reports Tab
        employee_reports_frame = ttk.Frame(report_notebook)
        report_notebook.add(employee_reports_frame, text="Employee Reports")
        self.create_employee_reports(employee_reports_frame)
        
        # Payroll Reports Tab
        payroll_reports_frame = ttk.Frame(report_notebook)
        report_notebook.add(payroll_reports_frame, text="Payroll Reports")
        self.create_payroll_reports(payroll_reports_frame)
        
        # Tax Reports Tab
        tax_reports_frame = ttk.Frame(report_notebook)
        report_notebook.add(tax_reports_frame, text="Tax Reports")
        self.create_tax_reports(tax_reports_frame)
        
        # Summary Reports Tab
        summary_reports_frame = ttk.Frame(report_notebook)
        report_notebook.add(summary_reports_frame, text="Summary Reports")
        self.create_summary_reports(summary_reports_frame)
    
    def create_settings_tab(self):
        """Create the settings tab"""
        settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(settings_frame, text="⚙️ Settings")
        
        # Title
        ttk.Label(settings_frame, text="Application Settings", font=("Arial", 16, "bold")).pack(pady=20)
        
        # Tax Rates Section
        tax_frame = ttk.LabelFrame(settings_frame, text="Tax Rates (%)")
        tax_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Federal Tax
        federal_frame = ttk.Frame(tax_frame)
        federal_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(federal_frame, text="Federal Tax:").pack(side=tk.LEFT)
        self.federal_tax_var = tk.DoubleVar(value=15.0)
        federal_spinbox = ttk.Spinbox(federal_frame, from_=0.0, to=50.0, increment=0.1, 
                                     textvariable=self.federal_tax_var, width=10)
        federal_spinbox.pack(side=tk.RIGHT)
        
        # State Tax
        state_frame = ttk.Frame(tax_frame)
        state_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(state_frame, text="State Tax:").pack(side=tk.LEFT)
        self.state_tax_var = tk.DoubleVar(value=5.0)
        state_spinbox = ttk.Spinbox(state_frame, from_=0.0, to=20.0, increment=0.1, 
                                   textvariable=self.state_tax_var, width=10)
        state_spinbox.pack(side=tk.RIGHT)
        
        # Social Security Tax (fixed)
        ss_frame = ttk.Frame(tax_frame)
        ss_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(ss_frame, text="Social Security Tax:").pack(side=tk.LEFT)
        ttk.Label(ss_frame, text="6.2% (Fixed)", foreground="gray").pack(side=tk.RIGHT)
        
        # Medicare Tax (fixed)
        medicare_frame = ttk.Frame(tax_frame)
        medicare_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(medicare_frame, text="Medicare Tax:").pack(side=tk.LEFT)
        ttk.Label(medicare_frame, text="1.45% (Fixed)", foreground="gray").pack(side=tk.RIGHT)
        
        # Advanced Tax Features Section
        advanced_tax_frame = ttk.LabelFrame(settings_frame, text="Advanced Tax Features")
        advanced_tax_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Meals Tax (for restaurants)
        meals_frame = ttk.Frame(advanced_tax_frame)
        meals_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(meals_frame, text="Meals Tax:").pack(side=tk.LEFT)
        self.meals_tax_var = tk.DoubleVar(value=0.0)
        meals_spinbox = ttk.Spinbox(meals_frame, from_=0.0, to=10.0, increment=0.1, 
                                   textvariable=self.meals_tax_var, width=10)
        meals_spinbox.pack(side=tk.RIGHT)
        ttk.Label(meals_frame, text="(for restaurant/hospitality businesses)", 
                 foreground="gray", font=("Arial", 8)).pack(side=tk.RIGHT, padx=(5, 10))
        
        # Service Charge Tax
        service_frame = ttk.Frame(advanced_tax_frame)
        service_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(service_frame, text="Service Charge Tax:").pack(side=tk.LEFT)
        self.service_tax_var = tk.DoubleVar(value=0.0)
        service_spinbox = ttk.Spinbox(service_frame, from_=0.0, to=10.0, increment=0.1, 
                                     textvariable=self.service_tax_var, width=10)
        service_spinbox.pack(side=tk.RIGHT)
        ttk.Label(service_frame, text="(for service charges)", 
                 foreground="gray", font=("Arial", 8)).pack(side=tk.RIGHT, padx=(5, 10))
        
        # PFML Tax (Paid Family Medical Leave)
        pfml_frame = ttk.Frame(advanced_tax_frame)
        pfml_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(pfml_frame, text="PFML Tax:").pack(side=tk.LEFT)
        self.pfml_tax_var = tk.DoubleVar(value=0.75)
        pfml_spinbox = ttk.Spinbox(pfml_frame, from_=0.0, to=5.0, increment=0.01, 
                                  textvariable=self.pfml_tax_var, width=10)
        pfml_spinbox.pack(side=tk.RIGHT)
        ttk.Label(pfml_frame, text="(Paid Family Medical Leave)", 
                 foreground="gray", font=("Arial", 8)).pack(side=tk.RIGHT, padx=(5, 10))
        
        # Business Type Selection
        business_type_frame = ttk.Frame(advanced_tax_frame)
        business_type_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(business_type_frame, text="Business Type:").pack(side=tk.LEFT)
        self.business_type_var = tk.StringVar(value="General")
        business_type_combo = ttk.Combobox(business_type_frame, textvariable=self.business_type_var, 
                                         values=["General", "Restaurant", "Hospitality", "Retail", "Service"], 
                                         state="readonly", width=15)
        business_type_combo.pack(side=tk.RIGHT, padx=(10, 0))
        
        # Auto-configure taxes based on business type
        business_type_combo.bind("<<ComboboxSelected>>", self.on_business_type_changed)
        
        # Business Information Section
        business_frame = ttk.LabelFrame(settings_frame, text="Business Information")
        business_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Business Name
        name_frame = ttk.Frame(business_frame)
        name_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(name_frame, text="Business Name:").pack(side=tk.LEFT)
        self.business_name_var = tk.StringVar(value="My Business")
        name_entry = ttk.Entry(name_frame, textvariable=self.business_name_var)
        name_entry.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        
        # Default Pay Frequency
        freq_frame = ttk.Frame(business_frame)
        freq_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(freq_frame, text="Default Pay Frequency:").pack(side=tk.LEFT)
        self.pay_frequency_var = tk.StringVar(value="Weekly")
        freq_combo = ttk.Combobox(freq_frame, textvariable=self.pay_frequency_var, 
                                 values=["Weekly", "Bi-weekly", "Semi-monthly", "Monthly"], 
                                 state="readonly", width=15)
        freq_combo.pack(side=tk.RIGHT, padx=(10, 0))
        
        # Buttons
        button_frame = ttk.Frame(settings_frame)
        button_frame.pack(fill=tk.X, padx=20, pady=20)
        
        ttk.Button(button_frame, text="Save Settings", command=self.save_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Load Defaults", command=self.load_default_settings).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Reset Settings", command=self.reset_settings).pack(side=tk.LEFT, padx=5)
        
        # Data Management Section
        data_frame = ttk.LabelFrame(settings_frame, text="Data Management")
        data_frame.pack(fill=tk.X, padx=20, pady=10)
        
        data_button_frame = ttk.Frame(data_frame)
        data_button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(data_button_frame, text="Backup Database", command=self.backup_database).pack(side=tk.LEFT, padx=5)
        ttk.Button(data_button_frame, text="Restore Database", command=self.restore_database).pack(side=tk.LEFT, padx=5)
        ttk.Button(data_button_frame, text="Export All Data", command=self.export_all_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(data_button_frame, text="Reset Setup", command=self.reset_setup).pack(side=tk.LEFT, padx=5)
        
        # Load current settings
        self.load_settings()
    
    def load_settings(self):
        """Load settings from database or use defaults"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Create settings table if it doesn't exist
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            ''')
            
            # Load tax rates
            cursor.execute("SELECT value FROM settings WHERE key = 'federal_tax_rate'")
            result = cursor.fetchone()
            if result:
                self.federal_tax_var.set(float(result[0]))
            
            cursor.execute("SELECT value FROM settings WHERE key = 'state_tax_rate'")
            result = cursor.fetchone()
            if result:
                self.state_tax_var.set(float(result[0]))
            
            # Load business info
            cursor.execute("SELECT value FROM settings WHERE key = 'business_name'")
            result = cursor.fetchone()
            if result:
                self.business_name_var.set(result[0])
            
            cursor.execute("SELECT value FROM settings WHERE key = 'pay_frequency'")
            result = cursor.fetchone()
            if result:
                self.pay_frequency_var.set(result[0])
            
            # Load advanced tax settings
            cursor.execute("SELECT value FROM settings WHERE key = 'meals_tax_rate'")
            result = cursor.fetchone()
            if result:
                self.meals_tax_var.set(float(result[0]))
            
            cursor.execute("SELECT value FROM settings WHERE key = 'service_tax_rate'")
            result = cursor.fetchone()
            if result:
                self.service_tax_var.set(float(result[0]))
            
            cursor.execute("SELECT value FROM settings WHERE key = 'pfml_tax_rate'")
            result = cursor.fetchone()
            if result:
                self.pfml_tax_var.set(float(result[0]))
            
            cursor.execute("SELECT value FROM settings WHERE key = 'business_type'")
            result = cursor.fetchone()
            if result:
                self.business_type_var.set(result[0])
            
            conn.close()
            
        except Exception as e:
            print(f"Error loading settings: {e}")
    
    def save_settings(self):
        """Save current settings to database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Save tax rates
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("federal_tax_rate", str(self.federal_tax_var.get())))
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("state_tax_rate", str(self.state_tax_var.get())))
            
            # Save business info
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("business_name", self.business_name_var.get()))
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("pay_frequency", self.pay_frequency_var.get()))
            
            # Save advanced tax settings
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("meals_tax_rate", str(self.meals_tax_var.get())))
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("service_tax_rate", str(self.service_tax_var.get())))
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("pfml_tax_rate", str(self.pfml_tax_var.get())))
            cursor.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                          ("business_type", self.business_type_var.get()))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", "Settings saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save settings: {e}")
    
    def load_default_settings(self):
        """Load default settings"""
        self.federal_tax_var.set(15.0)
        self.state_tax_var.set(5.0)
        self.business_name_var.set("My Business")
        self.pay_frequency_var.set("Weekly")
        self.meals_tax_var.set(0.0)
        self.service_tax_var.set(0.0)
        self.pfml_tax_var.set(0.75)
        self.business_type_var.set("General")
        messagebox.showinfo("Info", "Default settings loaded. Click 'Save Settings' to apply.")
    
    def reset_settings(self):
        """Reset settings to defaults and save"""
        self.load_default_settings()
        self.save_settings()
    
    def get_tax_rates(self):
        """Get current tax rates for calculations"""
        return {
            'federal': self.federal_tax_var.get() / 100.0,
            'state': self.state_tax_var.get() / 100.0,
            'ss': 0.062,  # Fixed
            'medicare': 0.0145,  # Fixed
            'meals': self.meals_tax_var.get() / 100.0,
            'service': self.service_tax_var.get() / 100.0,
            'pfml': self.pfml_tax_var.get() / 100.0
        }
    
    def on_business_type_changed(self, event=None):
        """Auto-configure tax rates based on business type"""
        business_type = self.business_type_var.get()
        
        if business_type == "Restaurant":
            self.meals_tax_var.set(6.25)  # Typical meals tax rate
            self.service_tax_var.set(0.0)
            self.pfml_tax_var.set(0.75)
        elif business_type == "Hospitality":
            self.meals_tax_var.set(6.25)
            self.service_tax_var.set(0.0)
            self.pfml_tax_var.set(0.75)
        elif business_type == "Service":
            self.meals_tax_var.set(0.0)
            self.service_tax_var.set(2.0)  # Service charge tax
            self.pfml_tax_var.set(0.75)
        elif business_type == "Retail":
            self.meals_tax_var.set(0.0)
            self.service_tax_var.set(0.0)
            self.pfml_tax_var.set(0.75)
        else:  # General
            self.meals_tax_var.set(0.0)
            self.service_tax_var.set(0.0)
            self.pfml_tax_var.set(0.75)
        
        messagebox.showinfo("Tax Configuration", f"Tax rates auto-configured for {business_type} business type.")
    
    def create_employee_reports(self, parent):
        """Create employee reports section"""
        # Employee List Report
        ttk.Label(parent, text="Employee Reports", font=("Arial", 12, "bold")).pack(pady=10)
        
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Button(button_frame, text="Export Employee List (CSV)", 
                  command=self.export_employees).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export Employee List (Excel)", 
                  command=self.export_employees_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Employee Summary Report", 
                  command=self.export_employee_summary).pack(side=tk.LEFT, padx=5)
    
    def create_payroll_reports(self, parent):
        """Create payroll reports section"""
        ttk.Label(parent, text="Payroll Reports", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Date range selection
        date_frame = ttk.LabelFrame(parent, text="Date Range")
        date_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(date_frame, text="From:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.start_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-01"))
        ttk.Entry(date_frame, textvariable=self.start_date_var).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(date_frame, text="To:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.end_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.end_date_var).grid(row=0, column=3, padx=5, pady=5)
        
        # Report buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Button(button_frame, text="Payroll Summary (CSV)", 
                  command=self.export_payroll_summary_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Payroll Summary (Excel)", 
                  command=self.export_payroll_summary_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Detailed Payroll Report", 
                  command=self.export_detailed_payroll).pack(side=tk.LEFT, padx=5)
    
    def create_tax_reports(self, parent):
        """Create tax reports section"""
        ttk.Label(parent, text="Tax Reports", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Date range selection
        date_frame = ttk.LabelFrame(parent, text="Date Range")
        date_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(date_frame, text="From:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.tax_start_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-01"))
        ttk.Entry(date_frame, textvariable=self.tax_start_date_var).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(date_frame, text="To:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.tax_end_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(date_frame, textvariable=self.tax_end_date_var).grid(row=0, column=3, padx=5, pady=5)
        
        # Report buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Button(button_frame, text="Tax Summary Report", 
                  command=self.export_tax_summary).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Quarterly Tax Report", 
                  command=self.export_quarterly_taxes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Year-end Tax Summary", 
                  command=self.export_year_end_taxes).pack(side=tk.LEFT, padx=5)
    
    def create_summary_reports(self, parent):
        """Create summary reports section"""
        ttk.Label(parent, text="Summary Reports", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Report buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Button(button_frame, text="Business Summary Report", 
                  command=self.export_business_summary).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Monthly Payroll Summary", 
                  command=self.export_monthly_summary).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Employee Earnings Report", 
                  command=self.export_employee_earnings).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cost Analysis Report", 
                  command=self.export_cost_analysis).pack(side=tk.LEFT, padx=5)
    
    def load_employees(self):
        """Load employees from database with performance optimizations"""
        try:
            conn = self.get_optimized_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, position, hourly_rate FROM employees ORDER BY name")
            employees = cursor.fetchall()
            
            # Clear listbox
            self.employee_listbox.delete(0, tk.END)
            
            # Populate listbox and combobox
            employee_names = []
            for employee in employees:
                self.employee_listbox.insert(tk.END, f"{employee[1]} - {employee[2]}")
                employee_names.append(employee[1])
            
            self.payroll_employee_combo['values'] = employee_names
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {e}")
    
    def on_employee_select(self, event):
        """Handle employee selection"""
        try:
            selection = self.employee_listbox.curselection()
            if selection:
                index = selection[0]
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT name, position, hourly_rate, phone, email FROM employees ORDER BY name LIMIT 1 OFFSET ?", (index,))
                employee = cursor.fetchone()
                conn.close()
                
                if employee:
                    self.name_entry.delete(0, tk.END)
                    self.name_entry.insert(0, employee[0])
                    self.position_entry.delete(0, tk.END)
                    self.position_entry.insert(0, employee[1] or "")
                    self.rate_entry.delete(0, tk.END)
                    self.rate_entry.insert(0, str(employee[2]))
                    self.phone_entry.delete(0, tk.END)
                    self.phone_entry.insert(0, employee[3] or "")
                    self.email_entry.delete(0, tk.END)
                    self.email_entry.insert(0, employee[4] or "")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employee details: {e}")
    
    def add_employee(self):
        """Add a new employee with enhanced validation"""
        try:
            self.show_loading_indicator("Adding employee...")
            name = self.name_entry.get().strip()
            position = self.position_entry.get().strip()
            rate_str = self.rate_entry.get().strip()
            phone = self.phone_entry.get().strip()
            email = self.email_entry.get().strip()
            
            # Enhanced validation
            if not name:
                messagebox.showerror("Validation Error", "Employee name is required")
                self.name_entry.focus()
                return
            
            if len(name) > 100:
                messagebox.showerror("Validation Error", "Employee name must be 100 characters or less")
                self.name_entry.focus()
                return
            
            if not rate_str:
                messagebox.showerror("Validation Error", "Hourly rate is required")
                self.rate_entry.focus()
                return
            
            try:
                rate = float(rate_str)
            except ValueError:
                messagebox.showerror("Validation Error", "Hourly rate must be a valid number")
                self.rate_entry.focus()
                return
            
            if rate < 0.01 or rate > 1000.00:
                messagebox.showerror("Validation Error", "Hourly rate must be between $0.01 and $1000.00")
                self.rate_entry.focus()
                return
            
            # Email validation
            if email and not self.validate_email(email):
                messagebox.showerror("Validation Error", "Please enter a valid email address")
                self.email_entry.focus()
                return
            
            # Phone validation
            if phone and not self.validate_phone(phone):
                messagebox.showerror("Validation Error", "Please enter a valid phone number (e.g., 555-123-4567)")
                self.phone_entry.focus()
                return
            
            # Check for duplicate employee names
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM employees WHERE LOWER(name) = LOWER(?)", (name,))
            if cursor.fetchone()[0] > 0:
                conn.close()
                messagebox.showerror("Validation Error", f"An employee with the name '{name}' already exists")
                self.name_entry.focus()
                return
            
            cursor.execute("""
                INSERT INTO employees (name, position, hourly_rate, phone, email)
                VALUES (?, ?, ?, ?, ?)
            """, (name, position, rate, phone, email))
            conn.commit()
            conn.close()
            
            self.load_employees()
            self.clear_form()
            self.update_save_status("saved")
            self.show_success_animation(f"Employee '{name}' added successfully!")
            messagebox.showinfo("Success", f"Employee '{name}' added successfully!")
            
        except sqlite3.Error as e:
            messagebox.showerror("Database Error", f"Database error: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add employee: {e}")
    
    def validate_email(self, email):
        """Validate email address format"""
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    def validate_phone(self, phone):
        """Validate phone number format"""
        import re
        # Accept formats like: 555-123-4567, (555) 123-4567, 555.123.4567, 5551234567
        pattern = r'^\(?([0-9]{3})\)?[-. ]?([0-9]{3})[-. ]?([0-9]{4})$'
        return re.match(pattern, phone) is not None
    
    def validate_date(self, date_str):
        """Validate date format (YYYY-MM-DD)"""
        import re
        pattern = r'^\d{4}-\d{2}-\d{2}$'
        if not re.match(pattern, date_str):
            return False
        try:
            from datetime import datetime
            datetime.strptime(date_str, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    
    def update_employee(self):
        """Update selected employee"""
        # Implementation for updating employee
        messagebox.showinfo("Info", "Update functionality coming soon!")
    
    def delete_employee(self):
        """Delete selected employee"""
        # Implementation for deleting employee
        messagebox.showinfo("Info", "Delete functionality coming soon!")
    
    def clear_form(self):
        """Clear the employee form"""
        self.name_entry.delete(0, tk.END)
        self.position_entry.delete(0, tk.END)
        self.rate_entry.delete(0, tk.END)
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
    
    def calculate_payroll(self):
        """Calculate payroll for selected employee with enhanced validation"""
        try:
            employee_name = self.payroll_employee_var.get()
            hours_str = self.hours_entry.get().strip()
            rate_str = self.payroll_rate_entry.get().strip()
            
            if not employee_name:
                messagebox.showerror("Validation Error", "Please select an employee")
                return
            
            if not hours_str:
                messagebox.showerror("Validation Error", "Please enter hours worked")
                self.hours_entry.focus()
                return
            
            if not rate_str:
                messagebox.showerror("Validation Error", "Please enter hourly rate")
                self.payroll_rate_entry.focus()
                return
            
            try:
                hours = float(hours_str)
            except ValueError:
                messagebox.showerror("Validation Error", "Hours worked must be a valid number")
                self.hours_entry.focus()
                return
            
            try:
                rate = float(rate_str)
            except ValueError:
                messagebox.showerror("Validation Error", "Hourly rate must be a valid number")
                self.payroll_rate_entry.focus()
                return
            
            if hours < 0 or hours > 168:  # Max 168 hours per week (7 days * 24 hours)
                messagebox.showerror("Validation Error", "Hours worked must be between 0 and 168")
                self.hours_entry.focus()
                return
            
            if rate < 0.01 or rate > 1000.00:
                messagebox.showerror("Validation Error", "Hourly rate must be between $0.01 and $1000.00")
                self.payroll_rate_entry.focus()
                return
            
            gross_pay = hours * rate
            tax_rates = self.get_tax_rates()
            federal_tax = gross_pay * tax_rates['federal']
            ss_tax = gross_pay * tax_rates['ss']
            medicare_tax = gross_pay * tax_rates['medicare']
            state_tax = gross_pay * tax_rates['state']
            meals_tax = gross_pay * tax_rates['meals']
            service_tax = gross_pay * tax_rates['service']
            pfml_tax = gross_pay * tax_rates['pfml']
            total_taxes = federal_tax + ss_tax + medicare_tax + state_tax + meals_tax + service_tax + pfml_tax
            net_pay = gross_pay - total_taxes
            
            self.gross_pay_label.config(text=f"${gross_pay:.2f}")
            self.taxes_label.config(text=f"${total_taxes:.2f}")
            self.net_pay_label.config(text=f"${net_pay:.2f}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate payroll: {e}")
    
    def save_payroll(self):
        """Save payroll run with enhanced validation"""
        try:
            # Use the same validation as calculate_payroll
            employee_name = self.payroll_employee_var.get()
            hours_str = self.hours_entry.get().strip()
            rate_str = self.payroll_rate_entry.get().strip()
            
            if not employee_name:
                messagebox.showerror("Validation Error", "Please select an employee")
                return
            
            if not hours_str:
                messagebox.showerror("Validation Error", "Please enter hours worked")
                self.hours_entry.focus()
                return
            
            if not rate_str:
                messagebox.showerror("Validation Error", "Please enter hourly rate")
                self.payroll_rate_entry.focus()
                return
            
            try:
                hours = float(hours_str)
            except ValueError:
                messagebox.showerror("Validation Error", "Hours worked must be a valid number")
                self.hours_entry.focus()
                return
            
            try:
                rate = float(rate_str)
            except ValueError:
                messagebox.showerror("Validation Error", "Hourly rate must be a valid number")
                self.payroll_rate_entry.focus()
                return
            
            if hours < 0 or hours > 168:
                messagebox.showerror("Validation Error", "Hours worked must be between 0 and 168")
                self.hours_entry.focus()
                return
            
            if rate < 0.01 or rate > 1000.00:
                messagebox.showerror("Validation Error", "Hourly rate must be between $0.01 and $1000.00")
                self.payroll_rate_entry.focus()
                return
            
            # Get employee ID
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM employees WHERE name = ?", (employee_name,))
            emp_result = cursor.fetchone()
            
            if not emp_result:
                messagebox.showerror("Error", "Employee not found in database")
                conn.close()
                return
            
            emp_id = emp_result[0]
            
            # Calculate payroll using settings
            gross_pay = hours * rate
            tax_rates = self.get_tax_rates()
            federal_tax = gross_pay * tax_rates['federal']
            ss_tax = gross_pay * tax_rates['ss']
            medicare_tax = gross_pay * tax_rates['medicare']
            state_tax = gross_pay * tax_rates['state']
            meals_tax = gross_pay * tax_rates['meals']
            service_tax = gross_pay * tax_rates['service']
            pfml_tax = gross_pay * tax_rates['pfml']
            total_taxes = federal_tax + ss_tax + medicare_tax + state_tax + meals_tax + service_tax + pfml_tax
            net_pay = gross_pay - total_taxes
            
            # Save to database
            run_date = self.payroll_date_entry.get().strip() or datetime.now().strftime("%Y-%m-%d")
            cursor.execute('''
                INSERT INTO payroll_runs (employee_id, run_date, frequency, hours_worked, 
                                        gross_pay, federal_tax, ss_tax, medicare_tax, 
                                        state_tax, meals_tax, service_tax, pfml_tax,
                                        total_taxes, net_pay, bonus)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (emp_id, run_date, "Weekly", hours, gross_pay, federal_tax, 
                  ss_tax, medicare_tax, state_tax, meals_tax, service_tax, pfml_tax,
                  total_taxes, net_pay, 0))
            
            conn.commit()
            conn.close()
            
            self.update_save_status("saved")
            messagebox.showinfo("Success", f"Payroll saved successfully!\nEmployee: {employee_name}\nGross Pay: ${gross_pay:.2f}\nNet Pay: ${net_pay:.2f}")
            self.update_status(f"Payroll saved for {employee_name}")
            
            # Refresh payroll history
            self.load_payroll_history()
            
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numeric values for hours and rate")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save payroll: {e}")
    
    def load_payroll_history(self):
        """Load and display payroll history"""
        try:
            # Clear existing items
            for item in self.payroll_tree.get_children():
                self.payroll_tree.delete(item)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT pr.id, e.name, pr.run_date, pr.hours_worked, 
                       pr.gross_pay, pr.net_pay
                FROM payroll_runs pr
                JOIN employees e ON pr.employee_id = e.id
                ORDER BY pr.run_date DESC
                LIMIT 50
            ''')
            
            payroll_records = cursor.fetchall()
            conn.close()
            
            for record in payroll_records:
                self.payroll_tree.insert("", "end", values=record)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payroll history: {e}")
    
    def edit_selected_payroll(self):
        """Edit selected payroll record"""
        selection = self.payroll_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a payroll record to edit")
            return
        
        item = self.payroll_tree.item(selection[0])
        run_id = item['values'][0]
        employee_name = item['values'][1]
        
        # Load payroll data for editing
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT pr.hours_worked, pr.gross_pay, pr.federal_tax, 
                       pr.ss_tax, pr.medicare_tax, pr.state_tax, 
                       pr.total_taxes, pr.net_pay, pr.bonus, pr.run_date
                FROM payroll_runs pr
                WHERE pr.id = ?
            ''', (run_id,))
            
            record = cursor.fetchone()
            conn.close()
            
            if record:
                # Create edit dialog
                self.create_edit_dialog(run_id, employee_name, record)
            else:
                messagebox.showerror("Error", "Payroll record not found")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payroll record: {e}")
    
    def create_edit_dialog(self, run_id, employee_name, record):
        """Create dialog for editing payroll record"""
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Payroll - {employee_name}")
        dialog.geometry("400x500")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (500 // 2)
        dialog.geometry(f"400x500+{x}+{y}")
        
        # Form fields
        ttk.Label(dialog, text=f"Editing payroll for: {employee_name}", font=("Arial", 12, "bold")).pack(pady=10)
        
        # Date
        date_frame = ttk.Frame(dialog)
        date_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(date_frame, text="Date:").pack(side=tk.LEFT)
        date_entry = ttk.Entry(date_frame)
        date_entry.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        date_entry.insert(0, record[9])  # run_date
        
        # Hours worked
        hours_frame = ttk.Frame(dialog)
        hours_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(hours_frame, text="Hours Worked:").pack(side=tk.LEFT)
        hours_entry = ttk.Entry(hours_frame)
        hours_entry.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        hours_entry.insert(0, str(record[0]))  # hours_worked
        
        # Gross pay
        gross_frame = ttk.Frame(dialog)
        gross_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(gross_frame, text="Gross Pay:").pack(side=tk.LEFT)
        gross_entry = ttk.Entry(gross_frame)
        gross_entry.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        gross_entry.insert(0, str(record[1]))  # gross_pay
        
        # Bonus
        bonus_frame = ttk.Frame(dialog)
        bonus_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Label(bonus_frame, text="Bonus:").pack(side=tk.LEFT)
        bonus_entry = ttk.Entry(bonus_frame)
        bonus_entry.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10, 0))
        bonus_entry.insert(0, str(record[8]))  # bonus
        
        def save_changes():
            try:
                new_date = date_entry.get()
                new_hours = float(hours_entry.get())
                new_gross = float(gross_entry.get())
                new_bonus = float(bonus_entry.get() or 0)
                
                # Recalculate taxes and net pay using settings
                tax_rates = self.get_tax_rates()
                federal_tax = new_gross * tax_rates['federal']
                ss_tax = new_gross * tax_rates['ss']
                medicare_tax = new_gross * tax_rates['medicare']
                state_tax = new_gross * tax_rates['state']
                meals_tax = new_gross * tax_rates['meals']
                service_tax = new_gross * tax_rates['service']
                pfml_tax = new_gross * tax_rates['pfml']
                total_taxes = federal_tax + ss_tax + medicare_tax + state_tax + meals_tax + service_tax + pfml_tax
                net_pay = new_gross - total_taxes
                
                # Update database
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE payroll_runs 
                    SET run_date=?, hours_worked=?, gross_pay=?, federal_tax=?,
                        ss_tax=?, medicare_tax=?, state_tax=?, meals_tax=?,
                        service_tax=?, pfml_tax=?, total_taxes=?, net_pay=?, bonus=?
                    WHERE id=?
                ''', (new_date, new_hours, new_gross, federal_tax, ss_tax, 
                      medicare_tax, state_tax, meals_tax, service_tax, pfml_tax,
                      total_taxes, net_pay, new_bonus, run_id))
                
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Success", "Payroll record updated successfully!")
                dialog.destroy()
                self.load_payroll_history()
                
            except ValueError:
                messagebox.showerror("Error", "Please enter valid numeric values")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to update payroll record: {e}")
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(fill=tk.X, padx=10, pady=20)
        
        ttk.Button(button_frame, text="Save Changes", command=save_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
    
    def delete_selected_payroll(self):
        """Delete selected payroll record"""
        selection = self.payroll_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a payroll record to delete")
            return
        
        item = self.payroll_tree.item(selection[0])
        run_id = item['values'][0]
        employee_name = item['values'][1]
        run_date = item['values'][2]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete payroll record for {employee_name} on {run_date}?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM payroll_runs WHERE id = ?", (run_id,))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Success", "Payroll record deleted successfully!")
                self.load_payroll_history()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete payroll record: {e}")
    
    def export_employees(self):
        """Export employee list to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT name, position, hourly_rate, phone, email FROM employees ORDER BY name")
                employees = cursor.fetchall()
                conn.close()
                
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Name', 'Position', 'Hourly Rate', 'Phone', 'Email'])
                    writer.writerows(employees)
                
                messagebox.showinfo("Success", f"Employee list exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export employees: {e}")
    
    def export_employees_excel(self):
        """Export employee list to Excel"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if filename:
                import openpyxl
                from openpyxl.styles import Font, Alignment
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Employee List"
                
                # Headers
                headers = ['ID', 'Name', 'Position', 'Hourly Rate', 'Hire Date', 'Phone', 'Email']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                # Data
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("SELECT id, name, position, hourly_rate, hire_date, phone, email FROM employees ORDER BY name")
                employees = cursor.fetchall()
                conn.close()
                
                for row, employee in enumerate(employees, 2):
                    for col, value in enumerate(employee, 1):
                        ws.cell(row=row, column=col, value=value)
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filename)
                messagebox.showinfo("Success", f"Employee list exported to {filename}")
                
        except ImportError:
            messagebox.showerror("Error", "Excel export requires openpyxl library. Please install it with: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export employees to Excel: {e}")
    
    def export_employee_summary(self):
        """Export employee summary report"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Get employee data with payroll statistics
                cursor.execute('''
                    SELECT e.id, e.name, e.position, e.hourly_rate, e.hire_date,
                           COUNT(pr.id) as payroll_runs,
                           COALESCE(SUM(pr.hours_worked), 0) as total_hours,
                           COALESCE(SUM(pr.gross_pay), 0) as total_gross,
                           COALESCE(AVG(pr.gross_pay), 0) as avg_gross
                    FROM employees e
                    LEFT JOIN payroll_runs pr ON e.id = pr.employee_id
                    GROUP BY e.id, e.name, e.position, e.hourly_rate, e.hire_date
                    ORDER BY e.name
                ''')
                
                employees = cursor.fetchall()
                conn.close()
                
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['ID', 'Name', 'Position', 'Hourly Rate', 'Hire Date', 
                                   'Payroll Runs', 'Total Hours', 'Total Gross Pay', 'Average Gross Pay'])
                    writer.writerows(employees)
                
                messagebox.showinfo("Success", f"Employee summary report exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export employee summary: {e}")
    
    def export_payroll_summary_csv(self):
        """Export payroll summary to CSV with date validation"""
        try:
            start_date = self.start_date_var.get().strip()
            end_date = self.end_date_var.get().strip()
            
            # Validate dates
            if not self.validate_date(start_date):
                messagebox.showerror("Validation Error", "Please enter a valid start date (YYYY-MM-DD)")
                return
            
            if not self.validate_date(end_date):
                messagebox.showerror("Validation Error", "Please enter a valid end date (YYYY-MM-DD)")
                return
            
            if start_date > end_date:
                messagebox.showerror("Validation Error", "Start date must be before or equal to end date")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                cursor.execute('''
                    SELECT pr.id, e.name, pr.run_date, pr.frequency, pr.hours_worked,
                           pr.gross_pay, pr.federal_tax, pr.ss_tax, pr.medicare_tax, 
                           pr.state_tax, pr.total_taxes, pr.net_pay, pr.bonus
                    FROM payroll_runs pr
                    JOIN employees e ON pr.employee_id = e.id
                    WHERE pr.run_date BETWEEN ? AND ?
                    ORDER BY pr.run_date DESC, e.name
                ''', (start_date, end_date))
                
                payroll_data = cursor.fetchall()
                conn.close()
                
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['ID', 'Employee', 'Date', 'Frequency', 'Hours', 'Gross Pay',
                                   'Federal Tax', 'SS Tax', 'Medicare Tax', 'State Tax', 
                                   'Total Taxes', 'Net Pay', 'Bonus'])
                    writer.writerows(payroll_data)
                
                messagebox.showinfo("Success", f"Payroll summary exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export payroll summary: {e}")
    
    def export_payroll_summary_excel(self):
        """Export payroll summary to Excel with formatting"""
        try:
            start_date = self.start_date_var.get()
            end_date = self.end_date_var.get()
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if filename:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Payroll Summary {start_date} to {end_date}"
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                # Headers
                headers = ['ID', 'Employee', 'Date', 'Frequency', 'Hours', 'Gross Pay',
                          'Federal Tax', 'SS Tax', 'Medicare Tax', 'State Tax', 
                          'Total Taxes', 'Net Pay', 'Bonus']
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Data
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                cursor.execute('''
                    SELECT pr.id, e.name, pr.run_date, pr.frequency, pr.hours_worked,
                           pr.gross_pay, pr.federal_tax, pr.ss_tax, pr.medicare_tax, 
                           pr.state_tax, pr.total_taxes, pr.net_pay, pr.bonus
                    FROM payroll_runs pr
                    JOIN employees e ON pr.employee_id = e.id
                    WHERE pr.run_date BETWEEN ? AND ?
                    ORDER BY pr.run_date DESC, e.name
                ''', (start_date, end_date))
                
                payroll_data = cursor.fetchall()
                conn.close()
                
                for row, record in enumerate(payroll_data, 2):
                    for col, value in enumerate(record, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        if col >= 6:  # Format currency columns
                            cell.number_format = '$#,##0.00'
                
                # Add summary row
                summary_row = len(payroll_data) + 3
                ws.cell(row=summary_row, column=5, value="TOTALS:").font = Font(bold=True)
                
                # Calculate totals
                total_hours = sum(record[4] for record in payroll_data)
                total_gross = sum(record[5] for record in payroll_data)
                total_federal = sum(record[6] for record in payroll_data)
                total_ss = sum(record[7] for record in payroll_data)
                total_medicare = sum(record[8] for record in payroll_data)
                total_state = sum(record[9] for record in payroll_data)
                total_taxes = sum(record[10] for record in payroll_data)
                total_net = sum(record[11] for record in payroll_data)
                total_bonus = sum(record[12] for record in payroll_data)
                
                totals = [total_hours, total_gross, total_federal, total_ss, 
                         total_medicare, total_state, total_taxes, total_net, total_bonus]
                
                for col, total in enumerate(totals, 5):
                    cell = ws.cell(row=summary_row, column=col, value=total)
                    cell.font = Font(bold=True)
                    if col >= 6:
                        cell.number_format = '$#,##0.00'
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filename)
                messagebox.showinfo("Success", f"Payroll summary exported to {filename}")
                
        except ImportError:
            messagebox.showerror("Error", "Excel export requires openpyxl library. Please install it with: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export payroll summary to Excel: {e}")
    
    def export_tax_summary(self):
        """Export tax summary report"""
        try:
            start_date = self.tax_start_date_var.get()
            end_date = self.tax_end_date_var.get()
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Get tax summary by employee
                cursor.execute('''
                    SELECT e.name, 
                           SUM(pr.federal_tax) as total_federal,
                           SUM(pr.ss_tax) as total_ss,
                           SUM(pr.medicare_tax) as total_medicare,
                           SUM(pr.state_tax) as total_state,
                           SUM(pr.total_taxes) as total_all_taxes,
                           COUNT(pr.id) as payroll_runs
                    FROM payroll_runs pr
                    JOIN employees e ON pr.employee_id = e.id
                    WHERE pr.run_date BETWEEN ? AND ?
                    GROUP BY e.id, e.name
                    ORDER BY e.name
                ''', (start_date, end_date))
                
                tax_data = cursor.fetchall()
                
                # Get overall totals
                cursor.execute('''
                    SELECT SUM(pr.federal_tax) as total_federal,
                           SUM(pr.ss_tax) as total_ss,
                           SUM(pr.medicare_tax) as total_medicare,
                           SUM(pr.state_tax) as total_state,
                           SUM(pr.total_taxes) as total_all_taxes
                    FROM payroll_runs pr
                    WHERE pr.run_date BETWEEN ? AND ?
                ''', (start_date, end_date))
                
                overall_totals = cursor.fetchone()
                conn.close()
                
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Tax Summary Report', f'{start_date} to {end_date}'])
                    writer.writerow([])
                    writer.writerow(['Employee', 'Federal Tax', 'SS Tax', 'Medicare Tax', 
                                   'State Tax', 'Total Taxes', 'Payroll Runs'])
                    
                    for record in tax_data:
                        writer.writerow(record)
                    
                    writer.writerow([])
                    writer.writerow(['TOTALS', overall_totals[0], overall_totals[1], 
                                   overall_totals[2], overall_totals[3], overall_totals[4]])
                
                messagebox.showinfo("Success", f"Tax summary report exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export tax summary: {e}")
    
    def export_business_summary(self):
        """Export business summary report"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Get business statistics
                cursor.execute("SELECT COUNT(*) FROM employees")
                total_employees = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(*) FROM payroll_runs")
                total_payroll_runs = cursor.fetchone()[0]
                
                cursor.execute("SELECT SUM(gross_pay) FROM payroll_runs")
                total_gross_pay = cursor.fetchone()[0] or 0
                
                cursor.execute("SELECT SUM(total_taxes) FROM payroll_runs")
                total_taxes_paid = cursor.fetchone()[0] or 0
                
                cursor.execute("SELECT SUM(net_pay) FROM payroll_runs")
                total_net_pay = cursor.fetchone()[0] or 0
                
                # Get current settings
                business_name = self.business_name_var.get()
                federal_rate = self.federal_tax_var.get()
                state_rate = self.state_tax_var.get()
                
                conn.close()
                
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Business Summary Report'])
                    writer.writerow(['Generated:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                    writer.writerow([])
                    writer.writerow(['Business Information'])
                    writer.writerow(['Business Name:', business_name])
                    writer.writerow(['Federal Tax Rate:', f"{federal_rate}%"])
                    writer.writerow(['State Tax Rate:', f"{state_rate}%"])
                    writer.writerow(['Social Security Rate:', '6.2%'])
                    writer.writerow(['Medicare Rate:', '1.45%'])
                    writer.writerow([])
                    writer.writerow(['Payroll Statistics'])
                    writer.writerow(['Total Employees:', total_employees])
                    writer.writerow(['Total Payroll Runs:', total_payroll_runs])
                    writer.writerow(['Total Gross Pay:', f"${total_gross_pay:.2f}"])
                    writer.writerow(['Total Taxes Paid:', f"${total_taxes_paid:.2f}"])
                    writer.writerow(['Total Net Pay:', f"${total_net_pay:.2f}"])
                
                messagebox.showinfo("Success", f"Business summary report exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export business summary: {e}")
    
    # Placeholder methods for other reports
    def export_detailed_payroll(self):
        messagebox.showinfo("Info", "Detailed Payroll Report - Coming in next update!")
    
    def export_quarterly_taxes(self):
        messagebox.showinfo("Info", "Quarterly Tax Report - Coming in next update!")
    
    def export_year_end_taxes(self):
        messagebox.showinfo("Info", "Year-end Tax Summary - Coming in next update!")
    
    def export_monthly_summary(self):
        messagebox.showinfo("Info", "Monthly Payroll Summary - Coming in next update!")
    
    def export_employee_earnings(self):
        messagebox.showinfo("Info", "Employee Earnings Report - Coming in next update!")
    
    def export_cost_analysis(self):
        messagebox.showinfo("Info", "Cost Analysis Report - Coming in next update!")
    
    def backup_database(self):
        """Create a backup of the current database"""
        try:
            import shutil
            from datetime import datetime
            
            # Create backup filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"payroll_backup_{timestamp}.db"
            
            # Ask user for backup location
            filename = filedialog.asksaveasfilename(
                title="Save Database Backup",
                defaultextension=".db",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")],
                initialvalue=backup_filename
            )
            
            if filename:
                # Copy the database file
                shutil.copy2(self.db_path, filename)
                
                # Also create a backup in the profiles directory
                profiles_backup = os.path.join(os.path.dirname(self.db_path), f"backup_{timestamp}.db")
                shutil.copy2(self.db_path, profiles_backup)
                
                messagebox.showinfo("Backup Successful", 
                    f"Database backup created successfully!\n\nSaved to:\n{filename}\n\nLocal backup:\n{profiles_backup}")
                
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup: {e}")
    
    def restore_database(self):
        """Restore database from backup"""
        try:
            # Ask user to select backup file
            filename = filedialog.askopenfilename(
                title="Select Database Backup to Restore",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")]
            )
            
            if filename:
                # Confirm restoration
                if messagebox.askyesno("Confirm Restore", 
                    f"Are you sure you want to restore from backup?\n\nThis will replace ALL current data with the backup data.\n\nBackup file: {os.path.basename(filename)}\n\nThis action cannot be undone!"):
                    
                    # Create a backup of current database before restore
                    import shutil
                    from datetime import datetime
                    current_backup = f"current_backup_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                    shutil.copy2(self.db_path, os.path.join(os.path.dirname(self.db_path), current_backup))
                    
                    # Restore from backup
                    shutil.copy2(filename, self.db_path)
                    
                    messagebox.showinfo("Restore Successful", 
                        f"Database restored successfully from backup!\n\nRestored from: {os.path.basename(filename)}\n\nCurrent data backed up to: {current_backup}\n\nPlease restart the application to see changes.")
                    
                    # Refresh the application data
                    self.load_employees()
                    self.load_payroll_history()
                    
        except Exception as e:
            messagebox.showerror("Restore Error", f"Failed to restore backup: {e}")
    
    def export_all_data(self):
        """Export all application data to a comprehensive file"""
        try:
            from datetime import datetime
            
            # Ask user for export location
            filename = filedialog.asksaveasfilename(
                title="Export All Data",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialvalue=f"payroll_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    
                    # Export settings
                    writer.writerow(['=== APPLICATION SETTINGS ==='])
                    cursor.execute("SELECT key, value FROM settings")
                    settings = cursor.fetchall()
                    for setting in settings:
                        writer.writerow(['SETTING', setting[0], setting[1]])
                    
                    writer.writerow([])
                    
                    # Export employees
                    writer.writerow(['=== EMPLOYEES ==='])
                    writer.writerow(['ID', 'Name', 'Position', 'Hourly Rate', 'Hire Date', 'Phone', 'Email'])
                    cursor.execute("SELECT * FROM employees ORDER BY name")
                    employees = cursor.fetchall()
                    for emp in employees:
                        writer.writerow(emp)
                    
                    writer.writerow([])
                    
                    # Export payroll runs
                    writer.writerow(['=== PAYROLL RUNS ==='])
                    writer.writerow(['ID', 'Employee ID', 'Run Date', 'Frequency', 'Hours Worked', 'Gross Pay',
                                   'Federal Tax', 'SS Tax', 'Medicare Tax', 'State Tax', 'Meals Tax',
                                   'Service Tax', 'PFML Tax', 'Total Taxes', 'Net Pay', 'Bonus'])
                    cursor.execute("SELECT * FROM payroll_runs ORDER BY run_date DESC")
                    payroll_runs = cursor.fetchall()
                    for run in payroll_runs:
                        writer.writerow(run)
                    
                    writer.writerow([])
                    
                    # Export summary
                    writer.writerow(['=== EXPORT SUMMARY ==='])
                    writer.writerow(['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                    writer.writerow(['Total Employees', len(employees)])
                    writer.writerow(['Total Payroll Runs', len(payroll_runs)])
                    writer.writerow(['Total Settings', len(settings)])
                
                conn.close()
                
                messagebox.showinfo("Export Successful", 
                    f"All data exported successfully!\n\nExported to: {filename}\n\nIncludes:\n- {len(employees)} employees\n- {len(payroll_runs)} payroll runs\n- {len(settings)} settings")
                
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {e}")
    
    def reset_setup(self):
        """Reset the application setup"""
        try:
            if messagebox.askyesno("Reset Setup", 
                "Are you sure you want to reset the application setup?\n\nThis will:\n"
                "- Delete all business information\n"
                "- Delete all employee data\n"
                "- Delete all payroll records\n"
                "- Return to the setup wizard\n\nThis action cannot be undone!"):
                
                # Close current application
                self.root.destroy()
                
                # Delete setup config
                setup_file = os.path.join(PROFILE_DIR, 'setup_config.json')
                if os.path.exists(setup_file):
                    os.remove(setup_file)
                
                # Delete database files
                if os.path.exists(PROFILE_DIR):
                    for file in os.listdir(PROFILE_DIR):
                        if file.endswith('.db'):
                            os.remove(os.path.join(PROFILE_DIR, file))
                
                messagebox.showinfo("Setup Reset", 
                    "Application setup has been reset.\n\nPlease restart the application to run the setup wizard again.")
                
        except Exception as e:
            messagebox.showerror("Reset Error", f"Failed to reset setup: {e}")
    
    def create_help_tab(self):
        """Create the help and documentation tab"""
        help_frame = ttk.Frame(self.notebook)
        self.notebook.add(help_frame, text="❓ Help & Support")
        
        # Create notebook for different help sections
        help_notebook = ttk.Notebook(help_frame)
        help_notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Quick Start Guide
        quickstart_frame = ttk.Frame(help_notebook)
        help_notebook.add(quickstart_frame, text="Quick Start")
        self.create_quickstart_help(quickstart_frame)
        
        # User Guide
        userguide_frame = ttk.Frame(help_notebook)
        help_notebook.add(userguide_frame, text="User Guide")
        self.create_userguide_help(userguide_frame)
        
        # Troubleshooting
        troubleshooting_frame = ttk.Frame(help_notebook)
        help_notebook.add(troubleshooting_frame, text="Troubleshooting")
        self.create_troubleshooting_help(troubleshooting_frame)
        
        # About
        about_frame = ttk.Frame(help_notebook)
        help_notebook.add(about_frame, text="About")
        self.create_about_help(about_frame)
    
    def create_quickstart_help(self, parent):
        """Create quick start guide"""
        # Create scrollable text widget
        text_frame = ttk.Frame(parent)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        help_text = """
🏢 PROFESSIONAL PAYROLL SYSTEM - QUICK START GUIDE

📋 GETTING STARTED
1. Configure your business settings in the Settings tab
2. Add your employees in the Employees tab
3. Process payroll in the Payroll tab
4. Generate reports in the Reports tab

⚙️ STEP 1: CONFIGURE SETTINGS
• Go to Settings tab
• Set your business name and type
• Configure tax rates (Federal, State, etc.)
• Choose default pay frequency
• Click "Save Settings"

👥 STEP 2: ADD EMPLOYEES
• Go to Employees tab
• Click "Add Employee"
• Fill in required information:
  - Name (required)
  - Position
  - Hourly Rate (required)
  - Phone Number
  - Email Address
• Click "Add Employee"

💰 STEP 3: PROCESS PAYROLL
• Go to Payroll tab
• Select an employee from dropdown
• Enter hours worked
• Enter hourly rate (or use employee's default)
• Click "Calculate Payroll" to see breakdown
• Click "Save Payroll" to save the payroll run

📊 STEP 4: GENERATE REPORTS
• Go to Reports tab
• Choose report type:
  - Employee Reports: Export employee lists
  - Payroll Reports: Export payroll data by date range
  - Tax Reports: Export tax summaries
  - Summary Reports: Export business statistics
• Select date ranges as needed
• Click desired export button

🔧 ADVANCED FEATURES
• Edit/Delete payroll runs in Payroll tab
• Configure industry-specific taxes (meals tax, service tax)
• Auto-configure taxes by business type
• Export to Excel with professional formatting

💡 TIPS
• Always save settings after making changes
• Use date ranges to filter reports
• Check validation messages for input errors
• Backup your data regularly
        """
        
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def create_userguide_help(self, parent):
        """Create detailed user guide"""
        text_frame = ttk.Frame(parent)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        help_text = """
📖 DETAILED USER GUIDE

🏢 EMPLOYEE MANAGEMENT
Adding Employees:
• Required fields: Name, Hourly Rate
• Optional fields: Position, Phone, Email
• Validation: Name must be unique, rate must be $0.01-$1000.00
• Email format validation included
• Phone number accepts various formats

Viewing Employees:
• All employees listed in table format
• Click employee to view details
• Use "Refresh" to update list

💰 PAYROLL PROCESSING
Calculating Payroll:
• Select employee from dropdown
• Enter hours worked (0-168 hours)
• Enter hourly rate or use employee default
• System automatically calculates all taxes

Tax Calculations:
• Federal Tax: Configurable rate (default 15%)
• State Tax: Configurable rate (default 5%)
• Social Security: Fixed 6.2%
• Medicare: Fixed 1.45%
• Meals Tax: For restaurants (configurable)
• Service Tax: For service businesses (configurable)
• PFML Tax: Paid Family Medical Leave (default 0.75%)

Saving Payroll:
• All calculations saved to database
• Payroll history maintained
• Edit/delete capabilities available

📊 REPORTING SYSTEM
Employee Reports:
• Export employee list to CSV/Excel
• Employee summary with payroll statistics

Payroll Reports:
• Date range filtering available
• Export to CSV/Excel with formatting
• Detailed breakdown of all taxes

Tax Reports:
• Tax summary by employee
• Overall tax totals
• Date range filtering

Summary Reports:
• Business overview statistics
• Total payroll costs
• Employee count and statistics

⚙️ SETTINGS CONFIGURATION
Tax Rates:
• Federal Tax: 0-50%
• State Tax: 0-20%
• Meals Tax: 0-10% (for restaurants)
• Service Tax: 0-10% (for service businesses)
• PFML Tax: 0-5%

Business Types:
• General: Standard configuration
• Restaurant: Auto-sets meals tax
• Hospitality: Auto-sets meals tax
• Service: Auto-sets service charge tax
• Retail: Standard configuration

Business Information:
• Business name
• Default pay frequency
• All settings saved automatically

🔧 ADVANCED FEATURES
Payroll Management:
• Edit existing payroll runs
• Delete payroll records
• Payroll history viewing
• Automatic tax recalculation

Data Validation:
• Email format validation
• Phone number validation
• Date format validation
• Numeric range validation
• Duplicate name checking

Export Options:
• CSV format for simple data
• Excel format with formatting
• Professional styling
• Auto-sizing columns
• Summary totals

💾 DATA MANAGEMENT
Database:
• SQLite database for reliability
• Automatic backups
• Data integrity constraints
• Foreign key relationships

File Management:
• Profile-based data storage
• Export capabilities
• Backup and restore options
        """
        
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def create_troubleshooting_help(self, parent):
        """Create troubleshooting guide"""
        text_frame = ttk.Frame(parent)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        help_text = """
🔧 TROUBLESHOOTING GUIDE

❌ COMMON ISSUES AND SOLUTIONS

🚫 Application Won't Start
• Check if Python is installed correctly
• Try running as administrator
• Check Windows Defender settings
• Ensure all files are in the same folder
• Try using 'py' instead of 'python' command

💾 Database Errors
• Check write permissions in application folder
• Ensure sufficient disk space available
• Restart the application
• Check if database file is corrupted
• Restore from backup if available

📧 Validation Errors
• Email Format: Use format like user@example.com
• Phone Format: Use formats like 555-123-4567 or (555) 123-4567
• Date Format: Use YYYY-MM-DD format (e.g., 2025-01-27)
• Numeric Values: Enter valid numbers only
• Required Fields: Fill in all required fields

💰 Payroll Calculation Issues
• Check tax rates in Settings tab
• Ensure hours are between 0 and 168
• Verify hourly rate is between $0.01 and $1000.00
• Select employee before calculating
• Save settings after changing tax rates

📊 Export Problems
• Check if destination folder has write permissions
• Ensure sufficient disk space
• Close Excel if file is open
• Try different file name or location
• Install openpyxl for Excel export: pip install openpyxl

⚙️ Settings Issues
• Save settings after making changes
• Restart application after major setting changes
• Check business type configuration
• Verify tax rate ranges
• Use "Reset Settings" to restore defaults

🔄 Performance Issues
• Close other applications to free memory
• Restart the application periodically
• Check database size and cleanup old data
• Ensure sufficient RAM (4GB minimum)

📁 File Access Issues
• Run application as administrator
• Check file permissions
• Ensure antivirus isn't blocking files
• Try moving application to different folder

🆘 GETTING HELP
• Check this troubleshooting guide first
• Review error messages carefully
• Note down exact error text
• Check application logs if available
• Contact system administrator

🔍 DEBUGGING TIPS
• Test with minimal data first
• Save settings frequently
• Backup data before major changes
• Use default settings to test functionality
• Check Windows Event Viewer for system errors

📋 SYSTEM REQUIREMENTS
• Windows 10/11 (64-bit)
• Python 3.8+ or py launcher
• 4GB RAM minimum (8GB recommended)
• 100MB free disk space
• 1024x768 screen resolution minimum

🛠️ INSTALLATION ISSUES
• Ensure Python is properly installed
• Check PATH environment variable
• Try different Python installation
• Use virtual environment if needed
• Check for conflicting Python versions
        """
        
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def create_about_help(self, parent):
        """Create about section"""
        text_frame = ttk.Frame(parent)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap=tk.WORD, font=("Arial", 10))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        about_text = """
ℹ️ ABOUT PROFESSIONAL PAYROLL SYSTEM

📋 APPLICATION INFORMATION
Version: 2.0.0
Build Date: January 2025
Developer: AI Assistant
Platform: Python 3.13+ with Tkinter
Database: SQLite 3

🎯 PURPOSE
Professional Payroll System is a comprehensive payroll management 
application designed for small to medium businesses. It provides 
complete payroll processing capabilities with industry-specific 
tax calculations and professional reporting.

✨ KEY FEATURES
• Multi-profile support for different businesses
• Complete employee management system
• Advanced payroll calculations with multiple tax types
• Industry-specific tax configurations
• Professional reporting and export capabilities
• Comprehensive data validation
• User-friendly interface with help system

🏢 SUPPORTED BUSINESS TYPES
• General Businesses
• Restaurants and Food Service
• Hospitality and Hotels
• Service Businesses
• Retail Operations

💰 TAX CALCULATIONS
• Federal Income Tax (configurable)
• State Tax (configurable)
• Social Security Tax (6.2%)
• Medicare Tax (1.45%)
• Meals Tax (for restaurants)
• Service Charge Tax (for service businesses)
• PFML Tax (Paid Family Medical Leave)

📊 REPORTING CAPABILITIES
• Employee Reports (CSV/Excel)
• Payroll Reports with date filtering
• Tax Summary Reports
• Business Summary Reports
• Professional Excel formatting
• Custom date ranges

🔧 TECHNICAL SPECIFICATIONS
• Built with Python 3.13+
• Tkinter GUI framework
• SQLite database engine
• OpenPyXL for Excel export
• Cross-platform compatibility
• Standalone executable option

🛡️ DATA SECURITY
• Local database storage
• Automatic data validation
• Backup and restore capabilities
• Data integrity constraints
• Profile-based isolation

📞 SUPPORT INFORMATION
• Built-in help system
• Comprehensive user documentation
• Troubleshooting guides
• Validation and error handling
• Professional user interface

🎉 ACKNOWLEDGMENTS
This application was developed using modern Python technologies
and best practices for business applications. It provides a 
professional solution for payroll management needs.

📄 LICENSE
This software is provided for business use. Please ensure 
compliance with local payroll and tax regulations.

🔄 UPDATE INFORMATION
Version 2.0.0 includes:
• Enhanced reporting system
• Advanced tax features
• Improved data validation
• Comprehensive help system
• Industry-specific configurations
• Professional Excel export

Thank you for using Professional Payroll System!
        """
        
        text_widget.insert(tk.END, about_text)
        text_widget.config(state=tk.DISABLED)
    
    def run(self):
        """Run the application"""
        self.root.mainloop()

if __name__ == "__main__":
    try:
        app = SimplePayrollApp()
        app.run()
    except Exception as e:
        # Show error in a simple window
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Application Error", f"Failed to start application:\n{e}")
        root.destroy()
