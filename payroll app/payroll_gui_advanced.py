import tkinter as tk

from tkinter import ttk, messagebox, filedialog, simpledialog
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import csv
from datetime import datetime
import payroll_db
import atexit
import time

# Add this import at the top
try:
    from tkcalendar import Calendar
    TKCALENDAR_AVAILABLE = True
except ImportError:
    TKCALENDAR_AVAILABLE = False
    Calendar = None # Ensure Calendar is None if import fails

PROFILE_DIR = 'profiles'
if not os.path.exists(PROFILE_DIR):
    os.makedirs(PROFILE_DIR)

def select_user_profile():
    """Prompt the user to select or create a profile. Returns the profile name and db path."""
    root = tk.Tk()
    root.withdraw() # Hide the main Tkinter window

    # Create profile selection dialog
    dialog = tk.Toplevel(root)
    dialog.title("Select Profile")
    dialog.geometry("400x200")
    dialog.transient(root)
    dialog.grab_set()
    dialog.resizable(False, False)

    # Center the dialog
    dialog.update_idletasks()
    x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
    y = (dialog.winfo_screenheight() // 2) - (200 // 2)
    dialog.geometry(f"400x200+{x}+{y}")

    profile_name_var = tk.StringVar()
    db_path_var = tk.StringVar()
    selected_profile = {"name": None, "db_path": None}

    def on_select():
        profile_name = profile_listbox.get(profile_listbox.curselection())
        selected_profile["name"] = profile_name
        selected_profile["db_path"] = os.path.join(PROFILE_DIR, f"{profile_name}.db")
        dialog.destroy()

    def on_create():
        new_profile_name = simpledialog.askstring("Create Profile", "Enter new profile name:")
        if new_profile_name:
            # Basic validation
            if not new_profile_name.strip():
                messagebox.showerror("Error", "Profile name cannot be empty.")
                return
            if any(char in r'/\:*?"<>|' for char in new_profile_name):
                messagebox.showerror("Error", "Profile name contains invalid characters.")
                return
            if new_profile_name in [f.replace('.db', '') for f in os.listdir(PROFILE_DIR) if f.endswith('.db')]:
                messagebox.showerror("Error", "Profile name already exists.")
                return

            selected_profile["name"] = new_profile_name
            selected_profile["db_path"] = os.path.join(PROFILE_DIR, f"{new_profile_name}.db")
            # Initialize a new database for the profile
            payroll_db.set_db_name(selected_profile["db_path"])
            payroll_db.create_tables()
            dialog.destroy()

    def on_cancel():
        selected_profile["name"] = None
        selected_profile["db_path"] = None
        dialog.destroy()

    tk.Label(dialog, text="Select an existing profile or create a new one:").pack(pady=10)

    profile_listbox = tk.Listbox(dialog, height=5)
    profile_listbox.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)

    # Populate listbox with existing profiles
    existing_profiles = [f.replace('.db', '') for f in os.listdir(PROFILE_DIR) if f.endswith('.db')]
    for profile in existing_profiles:
        profile_listbox.insert(tk.END, profile)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    tk.Button(button_frame, text="Select", command=on_select, width=10).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Create New", command=on_create, width=10).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Cancel", command=on_cancel, width=10).pack(side=tk.LEFT, padx=5)

    root.wait_window(dialog) # Wait for the dialog to close

    if selected_profile["name"] is None:
        # User cancelled or closed the dialog
        return None, None
    return selected_profile["name"], selected_profile["db_path"]

class EmployeeManagementFrame(ttk.Frame):
    def __init__(self, parent, db_path):
        super().__init__(parent)
        self.db_path = db_path
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            self.create_employee_interface()
            self.load_employees()
            self.built = True
    
    def create_employee_interface(self):
        # Main container with two panels
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel - Employee List
        list_frame = ttk.LabelFrame(main_frame, text="Employees", padding="10")
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Employee tree view
        self.employee_tree = ttk.Treeview(list_frame, columns=("ID", "Name", "Position", "Rate"), show="headings", height=15)
        self.employee_tree.heading("ID", text="ID")
        self.employee_tree.heading("Name", text="Name")
        self.employee_tree.heading("Position", text="Position")
        self.employee_tree.heading("Rate", text="Hourly Rate")
        
        self.employee_tree.column("ID", width=50, anchor=tk.CENTER)
        self.employee_tree.column("Name", width=150)
        self.employee_tree.column("Position", width=120)
        self.employee_tree.column("Rate", width=100, anchor=tk.E)
        
        # Scrollbar for tree
        tree_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.employee_tree.yview)
        self.employee_tree.configure(yscrollcommand=tree_scroll.set)
        
        self.employee_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind selection event
        self.employee_tree.bind("<<TreeviewSelect>>", self.on_employee_select)
        
        # Right panel - Employee Form
        form_frame = ttk.LabelFrame(main_frame, text="Employee Details", padding="10")
        form_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
        
        # Form fields
        self.create_form_fields(form_frame)
        
        # Buttons
        button_frame = ttk.Frame(form_frame)
        button_frame.pack(pady=(20, 0), fill=tk.X)
        
        ttk.Button(button_frame, text="Add Employee", command=self.add_employee).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Update", command=self.update_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete", command=self.delete_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        # Add import section
        import_frame = ttk.LabelFrame(form_frame, text="Bulk Import", padding="10")
        import_frame.pack(pady=(15, 0), fill=tk.X)
        
        ttk.Label(import_frame, text="Import multiple employees from CSV file", 
                 font=("Arial", 9)).pack(anchor=tk.W)
        ttk.Label(import_frame, text="Expected columns: First Name, Last Name, Email, Phone Number, Job Descriptions, Wages", 
                 font=("Arial", 8), foreground="gray").pack(anchor=tk.W, pady=(2, 5))
        
        ttk.Button(import_frame, text="📥 Import from CSV Template", 
                  command=self.import_from_csv).pack(fill=tk.X)
    
    def create_form_fields(self, parent):
        self.entries = {}
        fields = [
            ("Name*:", "name"),
            ("Position*:", "position"),
            ("Hourly Rate*:", "hourly_rate"),
            ("Hire Date:", "hire_date"),
            ("Phone:", "phone"),
            ("Email:", "email"),
            ("Address:", "address")
        ]
        
        for label_text, field_name in fields:
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, pady=2)
            
            ttk.Label(row, text=label_text, width=12).pack(side=tk.LEFT)
            entry = ttk.Entry(row, width=25)
            entry.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            self.entries[field_name] = entry
        
        # Add placeholder text for date format
        ttk.Label(parent, text="Date format: YYYY-MM-DD", font=("Arial", 8), foreground="gray").pack(pady=(5, 0))
    
    def load_employees(self):
        # Clear existing items
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        
        # Load employees from database
        try:
            payroll_db.set_db_name(self.db_path)
            employees = payroll_db.get_all_employees(self.db_path)
            for emp in employees:
                self.employee_tree.insert("", tk.END, values=(emp[0], emp[1], emp[2], f"${emp[3]:.2f}"))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {e}")
    
    def on_employee_select(self, event):
        selection = self.employee_tree.selection()
        if selection:
            item = self.employee_tree.item(selection[0])
            emp_id = item['values'][0]
            
            try:
                employee_data = payroll_db.get_employee_by_id(emp_id, self.db_path)
                if employee_data:
                    self.populate_form(employee_data)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load employee details: {e}")
    
    def populate_form(self, employee_data):
        self.clear_form()
        self.entries["name"].insert(0, employee_data[1] or "")
        self.entries["position"].insert(0, employee_data[2] or "")
        self.entries["hourly_rate"].insert(0, str(employee_data[3]) if employee_data[3] else "")
        self.entries["hire_date"].insert(0, employee_data[4] or "")
        self.entries["phone"].insert(0, employee_data[7] or "")
        self.entries["email"].insert(0, employee_data[8] or "")
        self.entries["address"].insert(0, employee_data[6] or "")
    
    def clear_form(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)
    
    def add_employee(self):
        if not self.validate_form():
            return
        
        try:
            name = self.entries["name"].get().strip()
            position = self.entries["position"].get().strip()
            hourly_rate = float(self.entries["hourly_rate"].get())
            hire_date = self.entries["hire_date"].get().strip() or None
            phone = self.entries["phone"].get().strip() or None
            email = self.entries["email"].get().strip() or None
            address = self.entries["address"].get().strip() or None
            
            employee_id = payroll_db.add_employee(
                name, position, hourly_rate, hire_date, None,
                address, phone, email, None, None, None, None, None, self.db_path
            )
            
            if employee_id:
                messagebox.showinfo("Success", f"Employee '{name}' added successfully!")
                self.clear_form()
                self.load_employees()
            else:
                messagebox.showerror("Error", "Failed to add employee")
                
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid hourly rate")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add employee: {e}")
    
    def update_employee(self):
        selection = self.employee_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an employee to update")
            return
        
        if not self.validate_form():
            return
        
        try:
            emp_id = self.employee_tree.item(selection[0])['values'][0]
            name = self.entries["name"].get().strip()
            position = self.entries["position"].get().strip()
            hourly_rate = float(self.entries["hourly_rate"].get())
            hire_date = self.entries["hire_date"].get().strip() or None
            phone = self.entries["phone"].get().strip() or None
            email = self.entries["email"].get().strip() or None
            address = self.entries["address"].get().strip() or None
            
            success = payroll_db.update_employee(
                emp_id, name, position, hourly_rate, hire_date, None,
                address, phone, email, None, None, None, None, None, self.db_path
            )
            
            if success:
                messagebox.showinfo("Success", f"Employee '{name}' updated successfully!")
                self.clear_form()
                self.load_employees()
            else:
                messagebox.showerror("Error", "Failed to update employee")
                
        except ValueError:
            messagebox.showerror("Error", "Please enter a valid hourly rate")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update employee: {e}")
    
    def delete_employee(self):
        selection = self.employee_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an employee to delete")
            return
        
        item = self.employee_tree.item(selection[0])
        emp_id = item['values'][0]
        emp_name = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{emp_name}'?\n\nThis will also delete all associated payroll records."):
            try:
                success = payroll_db.delete_employee(emp_id, self.db_path)
                if success:
                    messagebox.showinfo("Success", f"Employee '{emp_name}' deleted successfully!")
                    self.clear_form()
                    self.load_employees()
                else:
                    messagebox.showerror("Error", "Failed to delete employee")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete employee: {e}")
    
    def import_from_csv(self):
        """Import employees from CSV file"""
        file_path = filedialog.askopenfilename(
            title="Select Employee CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            imported_count = 0
            skipped_count = 0
            errors = []
            
            with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                # Try to detect delimiter
                sample = csvfile.read(1024)
                csvfile.seek(0)
                delimiter = ',' if ',' in sample else ';'
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Extract data from CSV
                        first_name = row.get('First Name', '').strip()
                        last_name = row.get('Last Name', '').strip()
                        
                        # Skip system accounts and empty rows
                        if not first_name or not last_name or first_name.lower() in ['login', 'default']:
                            skipped_count += 1
                            continue
                        
                        # Combine names
                        full_name = f"{first_name} {last_name}".strip()
                        
                        # Get job description (position)
                        job_descriptions = row.get('Job Descriptions', '').strip()
                        position = job_descriptions.split(';')[0].strip() if job_descriptions else 'Employee'
                        
                        # Get wage (hourly rate)
                        wages = row.get('Wages', '0').strip()
                        try:
                            # Handle multiple wages separated by semicolon
                            wage_value = wages.split(';')[0].strip() if wages else '0'
                            hourly_rate = float(wage_value) if wage_value else 0.0
                        except ValueError:
                            hourly_rate = 0.0
                        
                        # Get other fields
                        email = row.get('Email', '').strip()
                        phone = row.get('Phone Number', '').strip()
                        employee_id = row.get('Employee ID', '').strip()
                        
                        # Skip if essential data is missing
                        if not full_name or not position:
                            skipped_count += 1
                            continue
                        
                        # Check if employee already exists
                        existing_employees = payroll_db.get_all_employees(self.db_path)
                        existing_names = [emp[1].lower() for emp in existing_employees]
                        if full_name.lower() in existing_names:
                            skipped_count += 1
                            continue
                        
                        # Add employee to database
                        new_employee_id = payroll_db.add_employee(
                            full_name, position, hourly_rate, None, None,
                            None, phone, email, None, None, None, None, None, self.db_path
                        )
                        
                        if new_employee_id:
                            imported_count += 1
                        else:
                            errors.append(f"Row {row_num}: Failed to add {full_name}")
                            
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        continue
            
            # Show results
            result_message = f"Import completed!\n\n"
            result_message += f"✅ Imported: {imported_count} employees\n"
            result_message += f"⏭️ Skipped: {skipped_count} employees (duplicates/system accounts)\n"
            
            if errors:
                result_message += f"❌ Errors: {len(errors)}\n\n"
                if len(errors) <= 5:
                    result_message += "Errors:\n" + "\n".join(errors)
                else:
                    result_message += f"First 5 errors:\n" + "\n".join(errors[:5])
                    result_message += f"\n... and {len(errors)-5} more"
            
            messagebox.showinfo("CSV Import Results", result_message)
            
            # Refresh the employee list
            if imported_count > 0:
                self.load_employees()
                
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to import CSV file:\n{str(e)}")

    def validate_form(self):
        name = self.entries["name"].get().strip()
        position = self.entries["position"].get().strip()
        hourly_rate = self.entries["hourly_rate"].get().strip()
        
        if not name:
            messagebox.showerror("Validation Error", "Name is required")
            return False
        
        if not position:
            messagebox.showerror("Validation Error", "Position is required")
            return False
        
        if not hourly_rate:
            messagebox.showerror("Validation Error", "Hourly rate is required")
            return False
        
        try:
            rate = float(hourly_rate)
            if rate <= 0:
                messagebox.showerror("Validation Error", "Hourly rate must be greater than 0")
                return False
        except ValueError:
            messagebox.showerror("Validation Error", "Please enter a valid hourly rate")
            return False
        
        # Validate date format if provided
        hire_date = self.entries["hire_date"].get().strip()
        if hire_date:
            try:
                datetime.strptime(hire_date, "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Validation Error", "Please enter hire date in YYYY-MM-DD format")
                return False
        
        return True

class PayrollCalculationFrame(ttk.Frame):
    def __init__(self, parent, db_path):
        super().__init__(parent)
        self.db_path = db_path
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            self.create_payroll_interface()
            self.load_employees()
            self.built = True
    
    def create_payroll_interface(self):
        # Main container
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel - Payroll Form
        form_frame = ttk.LabelFrame(main_frame, text="Payroll Entry", padding="10")
        form_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        
        # Employee selection
        ttk.Label(form_frame, text="Employee:").pack(anchor=tk.W)
        self.employee_var = tk.StringVar()
        self.employee_combo = ttk.Combobox(form_frame, textvariable=self.employee_var, state="readonly", width=30)
        self.employee_combo.pack(pady=(0, 10), fill=tk.X)
        self.employee_combo.bind("<<ComboboxSelected>>", self.on_employee_selected)
        
        # Pay period
        ttk.Label(form_frame, text="Pay Period:").pack(anchor=tk.W)
        period_frame = ttk.Frame(form_frame)
        period_frame.pack(pady=(0, 10), fill=tk.X)
        
        self.start_date = ttk.Entry(period_frame, width=15)
        self.start_date.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(period_frame, text="to").pack(side=tk.LEFT, padx=5)
        
        self.end_date = ttk.Entry(period_frame, width=15)
        self.end_date.pack(side=tk.LEFT, padx=(5, 0))
        
        # Date format hint
        ttk.Label(form_frame, text="Date format: YYYY-MM-DD (e.g., 2024-01-15)", 
                 font=("Arial", 8), foreground="gray").pack(pady=(0, 10))
        
        # Hours worked
        ttk.Label(form_frame, text="Regular Hours:").pack(anchor=tk.W)
        self.regular_hours = ttk.Entry(form_frame, width=30)
        self.regular_hours.pack(pady=(0, 5), fill=tk.X)
        self.regular_hours.bind("<KeyRelease>", self.calculate_totals)
        
        ttk.Label(form_frame, text="Overtime Hours:").pack(anchor=tk.W)
        self.overtime_hours = ttk.Entry(form_frame, width=30)
        self.overtime_hours.pack(pady=(0, 10), fill=tk.X)
        self.overtime_hours.bind("<KeyRelease>", self.calculate_totals)
        
        # Current hourly rate (read-only)
        ttk.Label(form_frame, text="Hourly Rate:").pack(anchor=tk.W)
        self.hourly_rate = ttk.Entry(form_frame, state="readonly", width=30)
        self.hourly_rate.pack(pady=(0, 10), fill=tk.X)
        
        # Bonus/Deductions
        ttk.Label(form_frame, text="Bonus:").pack(anchor=tk.W)
        self.bonus = ttk.Entry(form_frame, width=30)
        self.bonus.pack(pady=(0, 5), fill=tk.X)
        self.bonus.bind("<KeyRelease>", self.calculate_totals)
        
        ttk.Label(form_frame, text="Deductions:").pack(anchor=tk.W)
        self.deductions = ttk.Entry(form_frame, width=30)
        self.deductions.pack(pady=(0, 10), fill=tk.X)
        self.deductions.bind("<KeyRelease>", self.calculate_totals)
        
        # Calculation results (read-only)
        results_frame = ttk.LabelFrame(form_frame, text="Calculation Results", padding="10")
        results_frame.pack(pady=(10, 0), fill=tk.X)
        
        self.create_result_fields(results_frame)
        
        # Buttons
        button_frame = ttk.Frame(form_frame)
        button_frame.pack(pady=(15, 0), fill=tk.X)
        
        ttk.Button(button_frame, text="Calculate", command=self.calculate_totals).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="Save Payroll", command=self.save_payroll).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear", command=self.clear_form).pack(side=tk.LEFT, padx=(5, 0))
        
        # Right panel - Payroll History
        history_frame = ttk.LabelFrame(main_frame, text="Recent Payroll Records", padding="10")
        history_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # History tree
        self.history_tree = ttk.Treeview(history_frame, columns=("Employee", "Period", "Hours", "Gross", "Net"), show="headings", height=20)
        self.history_tree.heading("Employee", text="Employee")
        self.history_tree.heading("Period", text="Pay Period")
        self.history_tree.heading("Hours", text="Total Hours")
        self.history_tree.heading("Gross", text="Gross Pay")
        self.history_tree.heading("Net", text="Net Pay")
        
        self.history_tree.column("Employee", width=120)
        self.history_tree.column("Period", width=100)
        self.history_tree.column("Hours", width=80, anchor=tk.E)
        self.history_tree.column("Gross", width=80, anchor=tk.E)
        self.history_tree.column("Net", width=80, anchor=tk.E)
        
        # History scrollbar
        history_scroll = ttk.Scrollbar(history_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=history_scroll.set)
        
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        history_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # History buttons
        hist_button_frame = ttk.Frame(history_frame)
        hist_button_frame.pack(pady=(10, 0), fill=tk.X)
        
        ttk.Button(hist_button_frame, text="Refresh", command=self.load_payroll_history).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(hist_button_frame, text="Delete Selected", command=self.delete_payroll_record).pack(side=tk.LEFT)
    
    def create_result_fields(self, parent):
        self.result_entries = {}
        fields = [
            ("Gross Pay:", "gross_pay"),
            ("Federal Tax (15%):", "federal_tax"),
            ("Social Security (6.2%):", "ss_tax"),
            ("Medicare (1.45%):", "medicare_tax"),
            ("State Tax (5%):", "state_tax"),
            ("Total Taxes:", "total_taxes"),
            ("Net Pay:", "net_pay")
        ]
        
        for label_text, field_name in fields:
            row = ttk.Frame(parent)
            row.pack(fill=tk.X, pady=1)
            
            ttk.Label(row, text=label_text, width=18).pack(side=tk.LEFT)
            entry = ttk.Entry(row, state="readonly", width=15)
            entry.pack(side=tk.RIGHT)
            self.result_entries[field_name] = entry
    
    def load_employees(self):
        try:
            payroll_db.set_db_name(self.db_path)
            employees = payroll_db.get_all_employees(self.db_path)
            employee_list = [f"{emp[1]} (ID: {emp[0]})" for emp in employees]
            self.employee_combo['values'] = employee_list
            if employee_list:
                self.employee_combo.set(employee_list[0])
                self.on_employee_selected()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {e}")
    
    def on_employee_selected(self, event=None):
        selection = self.employee_var.get()
        if selection:
            try:
                emp_id = selection.split("ID: ")[1].split(")")[0]
                employee_data = payroll_db.get_employee_by_id(int(emp_id), self.db_path)
                if employee_data:
                    self.hourly_rate.config(state="normal")
                    self.hourly_rate.delete(0, tk.END)
                    self.hourly_rate.insert(0, f"${employee_data[3]:.2f}")
                    self.hourly_rate.config(state="readonly")
                    self.calculate_totals()
            except Exception as e:
                messagebox.showerror("Error", f"Could not load employee data: {str(e)}")
    
    def calculate_totals(self, event=None):
        try:
            # Get values
            regular_hours = float(self.regular_hours.get() or 0)
            overtime_hours = float(self.overtime_hours.get() or 0)
            bonus = float(self.bonus.get() or 0)
            deductions = float(self.deductions.get() or 0)
            
            # Get hourly rate
            rate_text = self.hourly_rate.get().replace("$", "")
            hourly_rate = float(rate_text) if rate_text else 0
            
            # Calculate gross pay (overtime is 1.5x regular rate)
            regular_pay = regular_hours * hourly_rate
            overtime_pay = overtime_hours * hourly_rate * 1.5
            gross_pay = regular_pay + overtime_pay + bonus
            
            # Calculate taxes
            federal_tax = gross_pay * 0.15
            ss_tax = gross_pay * 0.062
            medicare_tax = gross_pay * 0.0145
            state_tax = gross_pay * 0.05
            total_taxes = federal_tax + ss_tax + medicare_tax + state_tax
            
            # Calculate net pay
            net_pay = gross_pay - total_taxes - deductions
            
            # Update result fields
            results = {
                "gross_pay": f"${gross_pay:.2f}",
                "federal_tax": f"${federal_tax:.2f}",
                "ss_tax": f"${ss_tax:.2f}",
                "medicare_tax": f"${medicare_tax:.2f}",
                "state_tax": f"${state_tax:.2f}",
                "total_taxes": f"${total_taxes:.2f}",
                "net_pay": f"${net_pay:.2f}"
            }
            
            for field, value in results.items():
                entry = self.result_entries[field]
                entry.config(state="normal")
                entry.delete(0, tk.END)
                entry.insert(0, value)
                entry.config(state="readonly")
                
        except ValueError:
            # Clear results if invalid input
            for entry in self.result_entries.values():
                entry.config(state="normal")
                entry.delete(0, tk.END)
                entry.config(state="readonly")
    
    def save_payroll(self):
        if not self.validate_payroll_form():
            return
        
        try:
            # Get employee ID
            selection = self.employee_var.get()
            emp_id = int(selection.split("ID: ")[1].split(")")[0])
            
            # Get form values
            start_date = self.start_date.get()
            end_date = self.end_date.get()
            regular_hours = float(self.regular_hours.get() or 0)
            overtime_hours = float(self.overtime_hours.get() or 0)
            bonus = float(self.bonus.get() or 0)
            deductions = float(self.deductions.get() or 0)
            
            # Get calculated values
            gross_pay = float(self.result_entries["gross_pay"].get().replace("$", ""))
            federal_tax = float(self.result_entries["federal_tax"].get().replace("$", ""))
            ss_tax = float(self.result_entries["ss_tax"].get().replace("$", ""))
            medicare_tax = float(self.result_entries["medicare_tax"].get().replace("$", ""))
            state_tax = float(self.result_entries["state_tax"].get().replace("$", ""))
            total_taxes = float(self.result_entries["total_taxes"].get().replace("$", ""))
            net_pay = float(self.result_entries["net_pay"].get().replace("$", ""))
            
            total_hours = regular_hours + overtime_hours
            
            # Save to database
            run_id = payroll_db.add_payroll_run(
                emp_id, end_date, "Custom", total_hours, gross_pay,
                federal_tax, ss_tax, medicare_tax, state_tax, 0,  # PFML tax = 0 for now
                total_taxes, net_pay, bonus, self.db_path
            )
            
            if run_id:
                messagebox.showinfo("Success", "Payroll record saved successfully!")
                self.clear_form()
                self.load_payroll_history()
            else:
                messagebox.showerror("Error", "Failed to save payroll record")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save payroll: {e}")
    
    def clear_form(self):
        self.start_date.delete(0, tk.END)
        self.end_date.delete(0, tk.END)
        self.regular_hours.delete(0, tk.END)
        self.overtime_hours.delete(0, tk.END)
        self.bonus.delete(0, tk.END)
        self.deductions.delete(0, tk.END)
        
        for entry in self.result_entries.values():
            entry.config(state="normal")
            entry.delete(0, tk.END)
            entry.config(state="readonly")
    
    def validate_payroll_form(self):
        if not self.employee_var.get():
            messagebox.showerror("Validation Error", "Please select an employee")
            return False
        
        if not self.start_date.get() or not self.end_date.get():
            messagebox.showerror("Validation Error", "Please enter start and end dates")
            return False
        
        # Validate date format
        try:
            datetime.strptime(self.start_date.get(), "%Y-%m-%d")
            datetime.strptime(self.end_date.get(), "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Validation Error", "Please enter dates in YYYY-MM-DD format")
            return False
        
        if not self.regular_hours.get() and not self.overtime_hours.get():
            messagebox.showerror("Validation Error", "Please enter hours worked")
            return False
        
        return True
    
    def load_payroll_history(self):
        # Clear existing items
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        try:
            payroll_data = payroll_db.get_all_payroll_runs_with_employee_names(self.db_path)
            for record in payroll_data[-20:]:  # Show last 20 records
                self.history_tree.insert("", tk.END, values=(
                    record[14],  # employee name
                    record[2],   # date
                    f"{record[4]:.1f}",  # hours
                    f"${record[5]:.2f}",  # gross pay
                    f"${record[12]:.2f}"  # net pay
                ))
        except Exception as e:
            pass  # Silent error - don't interrupt user experience
    
    def delete_payroll_record(self):
        selection = self.history_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a payroll record to delete")
            return
        
        if messagebox.askyesno("Confirm Delete", "Are you sure you want to delete this payroll record?"):
            try:
                # Note: This is a simplified implementation
                # In a real app, you'd need to get the record ID and use it
                messagebox.showinfo("Info", "Delete functionality would be implemented here")
                self.load_payroll_history()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete record: {e}")

class PastPayrollsFrame(ttk.Frame):
    def __init__(self, parent, db_path):
        super().__init__(parent)
        self.db_path = db_path
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            self.create_payroll_history_interface()
            self.load_all_payrolls()
            self.built = True
    
    def create_payroll_history_interface(self):
        # Main container
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title and controls
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(header_frame, text="Payroll History", font=("Arial", 14, "bold")).pack(side=tk.LEFT)
        
        # Filter controls
        filter_frame = ttk.Frame(header_frame)
        filter_frame.pack(side=tk.RIGHT)
        
        ttk.Label(filter_frame, text="Filter by Employee:").pack(side=tk.LEFT, padx=(0, 5))
        self.filter_var = tk.StringVar()
        self.filter_combo = ttk.Combobox(filter_frame, textvariable=self.filter_var, width=20)
        self.filter_combo.pack(side=tk.LEFT, padx=(0, 5))
        self.filter_combo.bind("<<ComboboxSelected>>", self.apply_filter)
        
        ttk.Button(filter_frame, text="Show All", command=self.show_all).pack(side=tk.LEFT, padx=(5, 0))
        
        # Payroll records tree
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        self.payroll_tree = ttk.Treeview(tree_frame, columns=("ID", "Employee", "Date", "Hours", "Gross", "Taxes", "Net"), show="headings")
        
        # Configure columns
        self.payroll_tree.heading("ID", text="Run ID")
        self.payroll_tree.heading("Employee", text="Employee")
        self.payroll_tree.heading("Date", text="Pay Date")
        self.payroll_tree.heading("Hours", text="Hours")
        self.payroll_tree.heading("Gross", text="Gross Pay")
        self.payroll_tree.heading("Taxes", text="Total Taxes")
        self.payroll_tree.heading("Net", text="Net Pay")
        
        self.payroll_tree.column("ID", width=80, anchor=tk.CENTER)
        self.payroll_tree.column("Employee", width=150)
        self.payroll_tree.column("Date", width=100, anchor=tk.CENTER)
        self.payroll_tree.column("Hours", width=80, anchor=tk.E)
        self.payroll_tree.column("Gross", width=100, anchor=tk.E)
        self.payroll_tree.column("Taxes", width=100, anchor=tk.E)
        self.payroll_tree.column("Net", width=100, anchor=tk.E)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.payroll_tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.payroll_tree.xview)
        self.payroll_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.payroll_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Bottom frame with buttons and summary
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Action buttons
        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side=tk.LEFT)
        
        ttk.Button(button_frame, text="Refresh", command=self.load_all_payrolls).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(button_frame, text="View Details", command=self.view_details).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete Selected", command=self.delete_selected).pack(side=tk.LEFT, padx=(5, 0))
        
        # Summary frame
        summary_frame = ttk.LabelFrame(bottom_frame, text="Summary", padding="5")
        summary_frame.pack(side=tk.RIGHT)
        
        self.summary_label = ttk.Label(summary_frame, text="No records loaded")
        self.summary_label.pack()
    
    def load_all_payrolls(self):
        # Clear existing items
        for item in self.payroll_tree.get_children():
            self.payroll_tree.delete(item)
        
        try:
            # Load employees for filter
            employees = payroll_db.get_all_employees(self.db_path)
            employee_names = ["All Employees"] + [emp[1] for emp in employees]
            self.filter_combo['values'] = employee_names
            self.filter_combo.set("All Employees")
            
            # Load payroll data
            payroll_data = payroll_db.get_all_payroll_runs_with_employee_names(self.db_path)
            total_gross = 0
            total_taxes = 0
            total_net = 0
            
            for record in reversed(payroll_data):  # Show newest first
                self.payroll_tree.insert("", tk.END, values=(
                    record[0],   # run_id
                    record[14],  # employee name
                    record[2],   # date
                    f"{record[4]:.1f}",  # hours
                    f"${record[5]:.2f}",  # gross pay
                    f"${record[11]:.2f}",  # total taxes
                    f"${record[12]:.2f}"   # net pay
                ))
                total_gross += record[5]
                total_taxes += record[11]
                total_net += record[12]
            
            # Update summary
            record_count = len(payroll_data)
            self.summary_label.config(text=f"Records: {record_count} | Total Gross: ${total_gross:.2f} | Total Taxes: ${total_taxes:.2f} | Total Net: ${total_net:.2f}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payroll data: {e}")
    
    def apply_filter(self, event=None):
        selected_employee = self.filter_var.get()
        if selected_employee == "All Employees":
            self.show_all()
            return
        
        # Clear and reload with filter
        for item in self.payroll_tree.get_children():
            self.payroll_tree.delete(item)
        
        try:
            payroll_data = payroll_db.get_all_payroll_runs_with_employee_names(self.db_path)
            filtered_data = [record for record in payroll_data if record[14] == selected_employee]
            
            total_gross = 0
            total_taxes = 0
            total_net = 0
            
            for record in reversed(filtered_data):
                self.payroll_tree.insert("", tk.END, values=(
                    record[0],   # run_id
                    record[14],  # employee name
                    record[2],   # date
                    f"{record[4]:.1f}",  # hours
                    f"${record[5]:.2f}",  # gross pay
                    f"${record[11]:.2f}",  # total taxes
                    f"${record[12]:.2f}"   # net pay
                ))
                total_gross += record[5]
                total_taxes += record[11]
                total_net += record[12]
            
            # Update summary
            record_count = len(filtered_data)
            self.summary_label.config(text=f"{selected_employee}: {record_count} records | Gross: ${total_gross:.2f} | Taxes: ${total_taxes:.2f} | Net: ${total_net:.2f}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to filter payroll data: {e}")
    
    def show_all(self):
        self.filter_combo.set("All Employees")
        self.load_all_payrolls()
    
    def view_details(self):
        selection = self.payroll_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a payroll record to view details")
            return
        
        item = self.payroll_tree.item(selection[0])
        run_id = item['values'][0]
        
        try:
            record = payroll_db.get_payroll_run_details(run_id, self.db_path)
            if record:
                details = f"""Payroll Run Details (ID: {run_id})
                
Employee: {item['values'][1]}
Pay Date: {record[2]}
Hours Worked: {record[4]:.1f}
Gross Pay: ${record[5]:.2f}
Federal Tax: ${record[6]:.2f}
Social Security: ${record[7]:.2f}
Medicare: ${record[8]:.2f}
State Tax: ${record[9]:.2f}
Total Taxes: ${record[11]:.2f}
Net Pay: ${record[12]:.2f}
Bonus: ${record[13]:.2f}"""
                
                messagebox.showinfo("Payroll Details", details)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payroll details: {e}")
    
    def export_to_excel(self):
        try:
            # Get all visible items
            items = []
            for item in self.payroll_tree.get_children():
                values = self.payroll_tree.item(item)['values']
                items.append(values)
            
            if not items:
                messagebox.showwarning("No Data", "No payroll records to export")
                return
            
            # Create Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Payroll History"
            
            # Headers
            headers = ["Run ID", "Employee", "Pay Date", "Hours", "Gross Pay", "Total Taxes", "Net Pay"]
            ws.append(headers)
            
            # Data
            for item in items:
                ws.append(item)
            
            # Save file
            filename = f"payroll_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            messagebox.showinfo("Export Complete", f"Payroll history exported to {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {e}")
    
    def delete_selected(self):
        selection = self.payroll_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a payroll record to delete")
            return
        
        item = self.payroll_tree.item(selection[0])
        run_id = item['values'][0]
        employee_name = item['values'][1]
        
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete payroll record for {employee_name} (Run ID: {run_id})?"):
            try:
                success = payroll_db.delete_payroll_run(run_id, self.db_path)
                if success:
                    messagebox.showinfo("Success", "Payroll record deleted successfully!")
                    self.load_all_payrolls()
                else:
                    messagebox.showerror("Error", "Failed to delete payroll record")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete record: {e}")

class MealsTaxFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            label = ttk.Label(self, text="Meals Tax - Feature in development")
            label.pack(pady=20)
            self.built = True

class RestaurantTaxesFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            label = ttk.Label(self, text="Restaurant Taxes - Feature in development")
            label.pack(pady=20)
            self.built = True

class CalendarFrame(ttk.Frame):
    def __init__(self, parent, db_path):
        super().__init__(parent)
        self.db_path = db_path
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            label = ttk.Label(self, text="Calendar - Feature in development")
            label.pack(pady=20)
            self.built = True

class SettingsFrame(ttk.Frame):
    def __init__(self, parent, profile_name, db_path):
        super().__init__(parent)
        self.profile_name = profile_name
        self.db_path = db_path
        self.built = False
        
    def ensure_built(self):
        if not self.built:
            label = ttk.Label(self, text="Settings - Feature in development")
            label.pack(pady=20)
            self.built = True

class MainApp(tk.Tk):
    def __init__(self, profile_name, db_path):
        super().__init__()
        
        # Remove loading screen for faster startup
        # self.loading_screen = LoadingScreen(self)
        # self.withdraw()  # Hide main window until ready
        
        # Initialize app with loading feedback
        self.initialize_app(profile_name, db_path)
        
        # Close loading screen and show main window
        # self.loading_screen.close()
        # self.deiconify()  # Show main window
        
        # Set up auto-save
        atexit.register(self.auto_save)

    def initialize_app(self, profile_name, db_path):
        """Initialize the application with loading feedback."""
        # Set up main window
        self.title(f"Professional Payroll Management System - {profile_name}")
        self.geometry("1400x800")
        
        # Set window icon and make it look more professional
        try:
            self.state('zoomed')  # Maximize on Windows
        except:
            pass
        self.resizable(True, True)
        self.profile_name = profile_name
        self.db_path = db_path
        
        # Initialize database
        payroll_db.set_db_name(self.db_path)
        payroll_db.create_tables(self.db_path)
        
        # Create interface components
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Create frame objects
        self.employee_frame = EmployeeManagementFrame(self.notebook, self.db_path)
        self.payroll_frame = PayrollCalculationFrame(self.notebook, self.db_path)
        self.past_payrolls_frame = PastPayrollsFrame(self.notebook, self.db_path)
        self.meals_tax_frame = MealsTaxFrame(self.notebook)
        self.restaurant_taxes_frame = RestaurantTaxesFrame(self.notebook)
        self.calendar_frame = CalendarFrame(self.notebook, self.db_path)
        self.settings_frame = SettingsFrame(self.notebook, self.profile_name, self.db_path)
        
        # Add tabs to interface
        self.notebook.add(self.employee_frame, text="Employee Management")
        self.notebook.add(self.payroll_frame, text="Payroll Calculation")
        self.notebook.add(self.past_payrolls_frame, text="Past Payrolls")
        self.notebook.add(self.meals_tax_frame, text="Meals Tax")
        self.notebook.add(self.restaurant_taxes_frame, text="Restaurant Taxes")
        self.notebook.add(self.calendar_frame, text="Calendar")
        self.notebook.add(self.settings_frame, text="Settings")
        
        # Set up event bindings
        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed)
        
        # Build the first tab immediately since it's the default
        self.employee_frame.ensure_built()

    def on_tab_changed(self, event):
        """Handle tab change events for lazy loading."""
        try:
            # Get the currently selected tab
            current_tab = self.notebook.select()
            tab_text = self.notebook.tab(current_tab, "text")
            
            # Map tab names to frame objects and ensure they're built
            tab_frames = {
                "Employee Management": self.employee_frame,
                "Payroll Calculation": self.payroll_frame,
                "Past Payrolls": self.past_payrolls_frame,
                "Meals Tax": self.meals_tax_frame,
                "Restaurant Taxes": self.restaurant_taxes_frame,
                "Calendar": self.calendar_frame,
                "Settings": self.settings_frame
            }
            
            # Build the selected tab if it exists
            if tab_text in tab_frames:
                frame = tab_frames[tab_text]
                if hasattr(frame, 'ensure_built'):
                    frame.ensure_built()
        except Exception as e:
            print(f"Error in on_tab_changed: {e}")

    def create_widgets(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(pady=10, padx=10, expand=True, fill=tk.BOTH)

        self.employee_tab = ttk.Frame(self.notebook)
        self.payroll_tab = ttk.Frame(self.notebook)
        self.reports_tab = ttk.Frame(self.notebook)

        self.notebook.add(self.employee_tab, text="Employees")
        self.notebook.add(self.payroll_tab, text="Payroll")
        self.notebook.add(self.reports_tab, text="Reports")

        self.create_employee_tab()
        self.create_payroll_tab()
        self.create_reports_tab()

    def create_employee_tab(self):
        # Employee List Frame
        employee_list_frame = ttk.LabelFrame(self.employee_tab, text="Employee List", padding="10")
        employee_list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.employee_tree = ttk.Treeview(employee_list_frame, columns=("ID", "Name", "Position", "Hourly Rate"), show="headings")
        self.employee_tree.heading("ID", text="ID")
        self.employee_tree.heading("Name", text="Name")
        self.employee_tree.heading("Position", text="Position")
        self.employee_tree.heading("Hourly Rate", text="Hourly Rate")

        self.employee_tree.column("ID", width=50, anchor=tk.CENTER)
        self.employee_tree.column("Name", width=150)
        self.employee_tree.column("Position", width=100)
        self.employee_tree.column("Hourly Rate", width=100, anchor=tk.E)

        self.employee_tree.pack(fill=tk.BOTH, expand=True)

        employee_scroll = ttk.Scrollbar(employee_list_frame, orient="vertical", command=self.employee_tree.yview)
        employee_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.employee_tree.configure(yscrollcommand=employee_scroll.set)

        self.employee_tree.bind("<<TreeviewSelect>>", self.load_employee_details)

        # Employee Details Frame
        employee_details_frame = ttk.LabelFrame(self.employee_tab, text="Employee Details", padding="10")
        employee_details_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)

        form_fields = [
            ("Employee ID:", "id_entry"),
            ("Name:", "name_entry"),
            ("Position:", "position_entry"),
            ("Hourly Rate:", "hourly_rate_entry"),
            ("Hire Date:", "hire_date_entry"),
            ("Date of Birth:", "dob_entry"),
            ("Address:", "address_entry"),
            ("Phone:", "phone_entry"),
            ("Email:", "email_entry"),
            ("Emergency Contact:", "emergency_contact_entry"),
            ("Tax ID:", "tax_id_entry"),
            ("Bank Name:", "bank_name_entry"),
            ("Account Number:", "account_number_entry"),
            ("Routing Number:", "routing_number_entry")
        ]

        self.employee_entries = {}
        for i, (label_text, entry_name) in enumerate(form_fields):
            row_frame = ttk.Frame(employee_details_frame)
            row_frame.pack(fill=tk.X, pady=2)
            ttk.Label(row_frame, text=label_text, width=15).pack(side=tk.LEFT, padx=5)
            entry = ttk.Entry(row_frame)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            self.employee_entries[entry_name] = entry

            # Make ID and Hire Date/DOB read-only for now, or manage their input
            if entry_name == "id_entry":
                entry.config(state='readonly')

            if entry_name in ["hire_date_entry", "dob_entry"] and TKCALENDAR_AVAILABLE:
                date_button = ttk.Button(row_frame, text="...", command=lambda e=entry: self.open_calendar(e))
                date_button.pack(side=tk.LEFT, padx=2)
            elif entry_name in ["hire_date_entry", "dob_entry"] and not TKCALENDAR_AVAILABLE:
                ttk.Label(row_frame, text="(YYYY-MM-DD)", foreground="gray").pack(side=tk.LEFT, padx=2)


        button_frame = ttk.Frame(employee_details_frame)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Add New", command=self.add_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Update", command=self.update_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Delete", command=self.delete_employee).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Form", command=self.clear_employee_form).pack(side=tk.LEFT, padx=5)

    def create_payroll_tab(self):
        # Left side: Payroll Run Input
        payroll_input_frame = ttk.LabelFrame(self.payroll_tab, text="New Payroll Run", padding="10")
        payroll_input_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=10, pady=10)

        form_fields = [
            ("Employee Name:", "employee_name_payroll_entry"),
            ("Run Date:", "run_date_entry"),
            ("Frequency:", "frequency_combobox"),
            ("Hours Worked:", "hours_worked_entry"),
            ("Gross Pay:", "gross_pay_entry"),
            ("Federal Tax:", "federal_tax_entry"),
            ("Social Security Tax:", "ss_tax_entry"),
            ("Medicare Tax:", "medicare_tax_entry"),
            ("MA State Tax:", "ma_state_tax_entry"),
            ("PFML Tax:", "pfml_tax_entry"),
            ("Total Taxes:", "total_taxes_entry"),
            ("Net Pay:", "net_pay_entry"),
            ("Bonus:", "bonus_entry")
        ]

        self.payroll_entries = {}
        for i, (label_text, entry_name) in enumerate(form_fields):
            row_frame = ttk.Frame(payroll_input_frame)
            row_frame.pack(fill=tk.X, pady=2)
            ttk.Label(row_frame, text=label_text, width=20).pack(side=tk.LEFT, padx=5)

            if entry_name == "frequency_combobox":
                entry = ttk.Combobox(row_frame, values=["Weekly", "Bi-Weekly", "Semi-Monthly", "Monthly"])
                entry.set("Bi-Weekly") # Default value
            elif entry_name == "employee_name_payroll_entry":
                entry = ttk.Combobox(row_frame)
                entry.bind("<<ComboboxSelected>>", self.on_payroll_employee_select)
            else:
                entry = ttk.Entry(row_frame)

            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            self.payroll_entries[entry_name] = entry

            if entry_name == "run_date_entry" and TKCALENDAR_AVAILABLE:
                date_button = ttk.Button(row_frame, text="...", command=lambda e=entry: self.open_calendar(e))
                date_button.pack(side=tk.LEFT, padx=2)
            elif entry_name == "run_date_entry" and not TKCALENDAR_AVAILABLE:
                ttk.Label(row_frame, text="(YYYY-MM-DD)", foreground="gray").pack(side=tk.LEFT, padx=2)


        # Auto-calculate button
        ttk.Button(payroll_input_frame, text="Calculate Payroll", command=self.calculate_payroll).pack(pady=10)

        button_frame = ttk.Frame(payroll_input_frame)
        button_frame.pack(pady=5)
        ttk.Button(button_frame, text="Save Payroll", command=self.save_payroll_run).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Form", command=self.clear_payroll_form).pack(side=tk.LEFT, padx=5)

        # Right side: Payroll History
        payroll_history_frame = ttk.LabelFrame(self.payroll_tab, text="Payroll History", padding="10")
        payroll_history_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.payroll_history_tree = ttk.Treeview(payroll_history_frame, columns=("Run ID", "Employee Name", "Date", "Frequency", "Gross Pay", "Net Pay"), show="headings")
        self.payroll_history_tree.heading("Run ID", text="Run ID")
        self.payroll_history_tree.heading("Employee Name", text="Employee")
        self.payroll_history_tree.heading("Date", text="Date")
        self.payroll_history_tree.heading("Frequency", text="Frequency")
        self.payroll_history_tree.heading("Gross Pay", text="Gross Pay")
        self.payroll_history_tree.heading("Net Pay", text="Net Pay")

        self.payroll_history_tree.column("Run ID", width=50, anchor=tk.CENTER)
        self.payroll_history_tree.column("Employee Name", width=150)
        self.payroll_history_tree.column("Date", width=100)
        self.payroll_history_tree.column("Frequency", width=80)
        self.payroll_history_tree.column("Gross Pay", width=100, anchor=tk.E)
        self.payroll_history_tree.column("Net Pay", width=100, anchor=tk.E)

        self.payroll_history_tree.pack(fill=tk.BOTH, expand=True)

        payroll_history_scroll = ttk.Scrollbar(payroll_history_frame, orient="vertical", command=self.payroll_history_tree.yview)
        payroll_history_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.payroll_history_tree.configure(yscrollcommand=payroll_history_scroll.set)

        self.payroll_history_tree.bind("<<TreeviewSelect>>", self.load_payroll_run_details)

        payroll_history_button_frame = ttk.Frame(payroll_history_frame)
        payroll_history_button_frame.pack(pady=5)
        ttk.Button(payroll_history_button_frame, text="Delete Selected Run", command=self.delete_selected_payroll_run).pack(side=tk.LEFT, padx=5)
        ttk.Button(payroll_history_button_frame, text="Refresh History", command=self.load_payroll_history).pack(side=tk.LEFT, padx=5)

        self.load_payroll_history()
        self.populate_employee_combobox()

    def create_reports_tab(self):
        ttk.Label(self.reports_tab, text="Reports Tab Content Here", font=('Inter', 14)).pack(pady=20)
        ttk.Button(self.reports_tab, text="Generate Payroll Report (Excel)", command=self.generate_payroll_report).pack(pady=10)
        ttk.Button(self.reports_tab, text="Generate Employee Report (Excel)", command=self.generate_employee_report).pack(pady=10)

    def open_calendar(self, entry_widget):
        """Opens a calendar dialog and inserts selected date into the entry widget."""
        if not TKCALENDAR_AVAILABLE:
            messagebox.showerror("Error", "tkcalendar library is not installed. Please install it using 'pip install tkcalendar' for date picker functionality.")
            return

        def set_date():
            selected_date = cal.selection_get()
            entry_widget.config(state='normal')
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, selected_date.strftime("%Y-%m-%d"))
            entry_widget.config(state='readonly' if entry_widget == self.employee_entries["id_entry"] else 'normal') # Revert ID to readonly
            top.destroy()

        top = tk.Toplevel(self)
        top.title("Select Date")
        top.grab_set()

        cal = Calendar(top, selectmode='day',
                       date_pattern='y-mm-dd')
        cal.pack(pady=20)

        ttk.Button(top, text="Select", command=set_date).pack(pady=10)

    def load_employees(self):
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        employees = payroll_db.get_all_employees()
        for emp in employees:
            self.employee_tree.insert("", tk.END, iid=emp[0], values=(emp[0], emp[1], emp[2], f"{emp[3]:.2f}"))
        self.populate_employee_combobox()

    def load_employee_details(self, event=None):
        selected_item = self.employee_tree.focus()
        if not selected_item:
            return

        employee_id = self.employee_tree.item(selected_item, "iid")
        employee_data = payroll_db.get_employee_by_id(employee_id)

        if employee_data:
            # Unpack data directly, assuming the order matches the DB schema
            (id, name, position, hourly_rate, hire_date, dob, address, phone, email,
             emergency_contact, tax_id, bank_name, account_number, routing_number) = employee_data

            self.clear_employee_form() # Clear first to ensure no old data remains

            # Enable ID entry temporarily to insert value, then disable
            self.employee_entries["id_entry"].config(state='normal')
            self.employee_entries["id_entry"].insert(0, id)
            self.employee_entries["id_entry"].config(state='readonly')

            self.employee_entries["name_entry"].insert(0, name)
            self.employee_entries["position_entry"].insert(0, position)
            self.employee_entries["hourly_rate_entry"].insert(0, f"{hourly_rate:.2f}")
            self.employee_entries["hire_date_entry"].insert(0, hire_date if hire_date else "")
            self.employee_entries["dob_entry"].insert(0, dob if dob else "")
            self.employee_entries["address_entry"].insert(0, address if address else "")
            self.employee_entries["phone_entry"].insert(0, phone if phone else "")
            self.employee_entries["email_entry"].insert(0, email if email else "")
            self.employee_entries["emergency_contact_entry"].insert(0, emergency_contact if emergency_contact else "")
            self.employee_entries["tax_id_entry"].insert(0, tax_id if tax_id else "")
            self.employee_entries["bank_name_entry"].insert(0, bank_name if bank_name else "")
            self.employee_entries["account_number_entry"].insert(0, account_number if account_number else "")
            self.employee_entries["routing_number_entry"].insert(0, routing_number if routing_number else "")

    def add_employee(self):
        try:
            name = self.employee_entries["name_entry"].get().strip()
            position = self.employee_entries["position_entry"].get().strip()
            hourly_rate_str = self.employee_entries["hourly_rate_entry"].get().strip()
            hire_date = self.employee_entries["hire_date_entry"].get().strip()
            dob = self.employee_entries["dob_entry"].get().strip()
            address = self.employee_entries["address_entry"].get().strip()
            phone = self.employee_entries["phone_entry"].get().strip()
            email = self.employee_entries["email_entry"].get().strip()
            emergency_contact = self.employee_entries["emergency_contact_entry"].get().strip()
            tax_id = self.employee_entries["tax_id_entry"].get().strip()
            bank_name = self.employee_entries["bank_name_entry"].get().strip()
            account_number = self.employee_entries["account_number_entry"].get().strip()
            routing_number = self.employee_entries["routing_number_entry"].get().strip()

            if not name or not position or not hourly_rate_str:
                messagebox.showerror("Input Error", "Name, Position, and Hourly Rate are required.")
                return

            try:
                hourly_rate = float(hourly_rate_str)
                if hourly_rate <= 0:
                    messagebox.showerror("Input Error", "Hourly Rate must be a positive number.")
                    return
            except ValueError:
                messagebox.showerror("Input Error", "Hourly Rate must be a valid number.")
                return

            # Validate date formats if provided
            if hire_date and not self._is_valid_date(hire_date):
                messagebox.showerror("Input Error", "Hire Date must be in YYYY-MM-DD format.")
                return
            if dob and not self._is_valid_date(dob):
                messagebox.showerror("Input Error", "Date of Birth must be in YYYY-MM-DD format.")
                return

            employee_id = payroll_db.add_employee(
                name, position, hourly_rate, hire_date or None, dob or None,
                address or None, phone or None, email or None,
                emergency_contact or None, tax_id or None,
                bank_name or None, account_number or None, routing_number or None
            )
            if employee_id:
                messagebox.showinfo("Success", f"Employee '{name}' added with ID: {employee_id}")
                self.clear_employee_form()
                self.load_employees()
            else:
                messagebox.showerror("Error", "Failed to add employee.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            print(f"Error in add_employee: {e}")


    def update_employee(self):
        selected_item = self.employee_tree.focus()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an employee to update.")
            return

        try:
            employee_id = int(self.employee_entries["id_entry"].get()) # ID is readonly, so it should be valid
            name = self.employee_entries["name_entry"].get().strip()
            position = self.employee_entries["position_entry"].get().strip()
            hourly_rate_str = self.employee_entries["hourly_rate_entry"].get().strip()
            hire_date = self.employee_entries["hire_date_entry"].get().strip()
            dob = self.employee_entries["dob_entry"].get().strip()
            address = self.employee_entries["address_entry"].get().strip()
            phone = self.employee_entries["phone_entry"].get().strip()
            email = self.employee_entries["email_entry"].get().strip()
            emergency_contact = self.employee_entries["emergency_contact_entry"].get().strip()
            tax_id = self.employee_entries["tax_id_entry"].get().strip()
            bank_name = self.employee_entries["bank_name_entry"].get().strip()
            account_number = self.employee_entries["account_number_entry"].get().strip()
            routing_number = self.employee_entries["routing_number_entry"].get().strip()

            if not name or not position or not hourly_rate_str:
                messagebox.showerror("Input Error", "Name, Position, and Hourly Rate are required.")
                return

            try:
                hourly_rate = float(hourly_rate_str)
                if hourly_rate <= 0:
                    messagebox.showerror("Input Error", "Hourly Rate must be a positive number.")
                    return
            except ValueError:
                messagebox.showerror("Input Error", "Hourly Rate must be a valid number.")
                return

            # Validate date formats if provided
            if hire_date and not self._is_valid_date(hire_date):
                messagebox.showerror("Input Error", "Hire Date must be in YYYY-MM-DD format.")
                return
            if dob and not self._is_valid_date(dob):
                messagebox.showerror("Input Error", "Date of Birth must be in YYYY-MM-DD format.")
                return

            success = payroll_db.update_employee(
                employee_id, name, position, hourly_rate, hire_date or None, dob or None,
                address or None, phone or None, email or None,
                emergency_contact or None, tax_id or None,
                bank_name or None, account_number or None, routing_number or None
            )
            if success:
                messagebox.showinfo("Success", f"Employee '{name}' updated successfully.")
                self.clear_employee_form()
                self.load_employees()
            else:
                messagebox.showerror("Error", "Failed to update employee.")
        except ValueError:
            messagebox.showerror("Error", "Invalid Employee ID.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            print(f"Error in update_employee: {e}")

    def delete_employee(self):
        selected_item = self.employee_tree.focus()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an employee to delete.")
            return

        employee_id = int(self.employee_tree.item(selected_item, "iid"))
        employee_name = self.employee_tree.item(selected_item, "values")[1]

        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete employee '{employee_name}' (ID: {employee_id})? This will also delete all associated payroll runs."):
            success = payroll_db.delete_employee(employee_id)
            if success:
                messagebox.showinfo("Success", f"Employee '{employee_name}' deleted successfully.")
                self.clear_employee_form()
                self.load_employees()
                self.load_payroll_history() # Refresh payroll history as well
            else:
                messagebox.showerror("Error", "Failed to delete employee.")

    def clear_employee_form(self):
        for entry_name, entry_widget in self.employee_entries.items():
            entry_widget.config(state='normal') # Enable to clear
            entry_widget.delete(0, tk.END)
            if entry_name == "id_entry":
                entry_widget.config(state='readonly') # Disable ID again

    def _is_valid_date(self, date_str):
        """Helper to validate YYYY-MM-DD date format."""
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
            return True
        except ValueError:
            return False

    def populate_employee_combobox(self):
        employees = payroll_db.get_all_employees()
        employee_names = [f"{emp[1]} (ID: {emp[0]})" for emp in employees] # Format: Name (ID: X)
        self.payroll_entries["employee_name_payroll_entry"]['values'] = employee_names
        if employee_names:
            self.payroll_entries["employee_name_payroll_entry"].set(employee_names[0]) # Set default

    def on_payroll_employee_select(self, event=None):
        selected_text = self.payroll_entries["employee_name_payroll_entry"].get()
        if not selected_text:
            return
        try:
            # Extract ID from "Name (ID: X)" format
            employee_id = int(selected_text.split("(ID: ")[1][:-1])
            employee_data = payroll_db.get_employee_by_id(employee_id)
            if employee_data:
                hourly_rate = employee_data[3] # Assuming hourly rate is at index 3
                self.payroll_entries["gross_pay_entry"].delete(0, tk.END)
                self.payroll_entries["gross_pay_entry"].insert(0, f"{hourly_rate:.2f}")
                self.payroll_entries["hours_worked_entry"].delete(0, tk.END) # Clear hours worked for new selection
                self.payroll_entries["hours_worked_entry"].insert(0, "0.0") # Default to 0 hours
            else:
                messagebox.showwarning("Employee Not Found", "Selected employee details could not be loaded.")
        except (ValueError, IndexError):
            messagebox.showerror("Selection Error", "Invalid employee selection format.")
            self.payroll_entries["employee_name_payroll_entry"].set("") # Clear invalid selection

    def calculate_payroll(self):
        try:
            selected_employee_text = self.payroll_entries["employee_name_payroll_entry"].get()
            if not selected_employee_text:
                messagebox.showerror("Input Error", "Please select an employee.")
                return

            employee_id = int(selected_employee_text.split("(ID: ")[1][:-1])
            employee_data = payroll_db.get_employee_by_id(employee_id)
            if not employee_data:
                messagebox.showerror("Error", "Could not retrieve employee details for calculation.")
                return

            hourly_rate = employee_data[3]
            hours_worked = float(self.payroll_entries["hours_worked_entry"].get())
            bonus = float(self.payroll_entries["bonus_entry"].get() or 0.0)

            gross_pay = (hourly_rate * hours_worked) + bonus

            # Simplified tax calculation (example percentages)
            federal_tax_rate = 0.15
            ss_tax_rate = 0.062 # Social Security
            medicare_tax_rate = 0.0145 # Medicare
            ma_state_tax_rate = 0.05 # Massachusetts state tax (example)
            pfml_tax_rate = 0.0068 # MA PFML (example)

            federal_tax = gross_pay * federal_tax_rate
            ss_tax = gross_pay * ss_tax_rate
            medicare_tax = gross_pay * medicare_tax_rate
            ma_state_tax = gross_pay * ma_state_tax_rate
            pfml_tax = gross_pay * pfml_tax_rate

            total_taxes = federal_tax + ss_tax + medicare_tax + ma_state_tax + pfml_tax
            net_pay = gross_pay - total_taxes

            self.payroll_entries["gross_pay_entry"].delete(0, tk.END)
            self.payroll_entries["gross_pay_entry"].insert(0, f"{gross_pay:.2f}")
            self.payroll_entries["federal_tax_entry"].delete(0, tk.END)
            self.payroll_entries["federal_tax_entry"].insert(0, f"{federal_tax:.2f}")
            self.payroll_entries["ss_tax_entry"].delete(0, tk.END)
            self.payroll_entries["ss_tax_entry"].insert(0, f"{ss_tax:.2f}")
            self.payroll_entries["medicare_tax_entry"].delete(0, tk.END)
            self.payroll_entries["medicare_tax_entry"].insert(0, f"{medicare_tax:.2f}")
            self.payroll_entries["ma_state_tax_entry"].delete(0, tk.END)
            self.payroll_entries["ma_state_tax_entry"].insert(0, f"{ma_state_tax:.2f}")
            self.payroll_entries["pfml_tax_entry"].delete(0, tk.END)
            self.payroll_entries["pfml_tax_entry"].insert(0, f"{pfml_tax:.2f}")
            self.payroll_entries["total_taxes_entry"].delete(0, tk.END)
            self.payroll_entries["total_taxes_entry"].insert(0, f"{total_taxes:.2f}")
            self.payroll_entries["net_pay_entry"].delete(0, tk.END)
            self.payroll_entries["net_pay_entry"].insert(0, f"{net_pay:.2f}")
            self.payroll_entries["bonus_entry"].delete(0, tk.END)
            self.payroll_entries["bonus_entry"].insert(0, f"{bonus:.2f}") # Ensure bonus is reflected

        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numbers for Hours Worked and Bonus.")
        except Exception as e:
            messagebox.showerror("Calculation Error", f"An unexpected error occurred during payroll calculation: {e}")
            print(f"Error in calculate_payroll: {e}")

    def save_payroll_run(self):
        try:
            selected_employee_text = self.payroll_entries["employee_name_payroll_entry"].get()
            if not selected_employee_text:
                messagebox.showerror("Input Error", "Please select an employee.")
                return
            employee_id = int(selected_employee_text.split("(ID: ")[1][:-1])

            run_date = self.payroll_entries["run_date_entry"].get()
            frequency = self.payroll_entries["frequency_combobox"].get()
            hours_worked = float(self.payroll_entries["hours_worked_entry"].get() or 0.0)
            gross_pay = float(self.payroll_entries["gross_pay_entry"].get())
            federal_tax = float(self.payroll_entries["federal_tax_entry"].get() or 0.0)
            ss_tax = float(self.payroll_entries["ss_tax_entry"].get() or 0.0)
            medicare_tax = float(self.payroll_entries["medicare_tax_entry"].get() or 0.0)
            ma_state_tax = float(self.payroll_entries["ma_state_tax_entry"].get() or 0.0)
            pfml_tax = float(self.payroll_entries["pfml_tax_entry"].get() or 0.0)
            total_taxes = float(self.payroll_entries["total_taxes_entry"].get())
            net_pay = float(self.payroll_entries["net_pay_entry"].get())
            bonus = float(self.payroll_entries["bonus_entry"].get() or 0.0)

            if not self._is_valid_date(run_date):
                messagebox.showerror("Input Error", "Run Date must be in YYYY-MM-DD format.")
                return

            run_id = payroll_db.add_payroll_run(
                employee_id, run_date, frequency, hours_worked, gross_pay,
                federal_tax, ss_tax, medicare_tax, ma_state_tax, pfml_tax,
                total_taxes, net_pay, bonus
            )
            if run_id:
                messagebox.showinfo("Success", f"Payroll run saved successfully (Run ID: {run_id}).")
                self.clear_payroll_form()
                self.load_payroll_history()
            else:
                messagebox.showerror("Error", "Failed to save payroll run.")
        except ValueError:
            messagebox.showerror("Input Error", "Please ensure all numeric fields are filled correctly.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")
            print(f"Error in save_payroll_run: {e}")

    def clear_payroll_form(self):
        # Clear all entry widgets in the payroll tab
        for entry_widget in self.payroll_entries.values():
            if isinstance(entry_widget, (ttk.Entry, ttk.Combobox)):
                entry_widget.delete(0, tk.END)
        # Reset combobox default if needed
        self.payroll_entries["frequency_combobox"].set("Bi-Weekly")
        self.payroll_entries["hours_worked_entry"].insert(0, "0.0")
        self.payroll_entries["bonus_entry"].insert(0, "0.0")
        self.populate_employee_combobox() # Re-populate and set default employee

    def load_payroll_history(self):
        for item in self.payroll_history_tree.get_children():
            self.payroll_history_tree.delete(item)

        history_data = payroll_db.get_all_payroll_runs_with_employee_names(self.db_path)
        for run in history_data:
            # run: (run_id, employee_id, run_date, frequency, hours_worked, gross_pay, federal_tax, ss_tax, medicare_tax, ma_state_tax, pfml_tax, total_taxes, net_pay, bonus, employee_name)
            self.payroll_history_tree.insert("", tk.END, iid=run[0],
                                             values=(run[0], run[14], run[2], run[3], f"{run[5]:.2f}", f"{run[12]:.2f}"))

    def load_payroll_run_details(self, event=None):
        selected_item = self.payroll_history_tree.focus()
        if not selected_item:
            return

        run_id = self.payroll_history_tree.item(selected_item, "iid")
        payroll_run_data = payroll_db.get_payroll_run_details(run_id)

        if payroll_run_data:
            # Unpack data: (run_id, employee_id, run_date, frequency, hours_worked, gross_pay, federal_tax, ss_tax, medicare_tax, ma_state_tax, pfml_tax, total_taxes, net_pay, bonus)
            (run_id, employee_id, run_date, frequency, hours_worked, gross_pay,
             federal_tax, ss_tax, medicare_tax, ma_state_tax, pfml_tax,
             total_taxes, net_pay, bonus) = payroll_run_data

            employee_name = payroll_db.get_employee_by_id(employee_id)[1] # Get name for display

            self.clear_payroll_form() # Clear form before loading

            self.payroll_entries["employee_name_payroll_entry"].set(f"{employee_name} (ID: {employee_id})")
            self.payroll_entries["run_date_entry"].insert(0, run_date)
            self.payroll_entries["frequency_combobox"].set(frequency)
            self.payroll_entries["hours_worked_entry"].insert(0, f"{hours_worked:.2f}")
            self.payroll_entries["gross_pay_entry"].insert(0, f"{gross_pay:.2f}")
            self.payroll_entries["federal_tax_entry"].insert(0, f"{federal_tax:.2f}")
            self.payroll_entries["ss_tax_entry"].insert(0, f"{ss_tax:.2f}")
            self.payroll_entries["medicare_tax_entry"].insert(0, f"{medicare_tax:.2f}")
            self.payroll_entries["ma_state_tax_entry"].insert(0, f"{ma_state_tax:.2f}")
            self.payroll_entries["pfml_tax_entry"].insert(0, f"{pfml_tax:.2f}")
            self.payroll_entries["total_taxes_entry"].insert(0, f"{total_taxes:.2f}")
            self.payroll_entries["net_pay_entry"].insert(0, f"{net_pay:.2f}")
            self.payroll_entries["bonus_entry"].insert(0, f"{bonus:.2f}")
        else:
            messagebox.showwarning("Not Found", "Payroll run details could not be loaded.")

    def delete_selected_payroll_run(self):
        selected_item = self.payroll_history_tree.focus()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a payroll run to delete.")
            return

        run_id = self.payroll_history_tree.item(selected_item, "iid")
        employee_name = self.payroll_history_tree.item(selected_item, "values")[1]
        run_date = self.payroll_history_tree.item(selected_item, "values")[2]

        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete payroll run for '{employee_name}' on {run_date} (Run ID: {run_id})?"):
            success = payroll_db.delete_payroll_run(run_id)
            if success:
                messagebox.showinfo("Success", "Payroll run deleted successfully.")
                self.load_payroll_history()
                self.clear_payroll_form()
            else:
                messagebox.showerror("Error", "Failed to delete payroll run.")

    def generate_payroll_report(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                title="Save Payroll Report")
        if not file_path:
            return

        try:
            payroll_data = payroll_db.get_all_payroll_runs_with_employee_names(self.db_path)
            if not payroll_data:
                messagebox.showinfo("No Data", "No payroll data to generate report.")
                return

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Payroll Report"

            # Headers
            headers = ["Run ID", "Employee ID", "Employee Name", "Run Date", "Frequency", "Hours Worked",
                       "Gross Pay", "Federal Tax", "SS Tax", "Medicare Tax", "MA State Tax", "PFML Tax",
                       "Total Taxes", "Net Pay", "Bonus"]
            sheet.append(headers)

            # Apply header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, cell in enumerate(sheet[1]):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = 15 # Default width

            # Data rows
            for row_idx, run in enumerate(payroll_data):
                # run: (run_id, employee_id, run_date, frequency, hours_worked, gross_pay, federal_tax, ss_tax, medicare_tax, ma_state_tax, pfml_tax, total_taxes, net_pay, bonus, employee_name)
                # Note: employee_name is the last element from the join in get_all_payroll_runs_with_employee_names
                sheet.append([
                    run[0], run[1], run[14], run[2], run[3], run[4],
                    run[5], run[6], run[7], run[8], run[9], run[10],
                    run[11], run[12], run[13]
                ])
                # Format currency columns
                for col_num in [7, 8, 9, 10, 11, 12, 13, 14, 15]: # Gross Pay to Bonus
                    sheet.cell(row=row_idx + 2, column=col_num).number_format = '"$"#,##0.00'

            # Auto-size columns for better readability
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Report Generated", f"Payroll report saved to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate payroll report: {e}")
            print(f"Error generating payroll report: {e}")

    def generate_employee_report(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                filetypes=[("Excel files", "*.xlsx")],
                                                title="Save Employee Report")
        if not file_path:
            return

        try:
            employee_data = payroll_db.get_all_employees()
            if not employee_data:
                messagebox.showinfo("No Data", "No employee data to generate report.")
                return

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Employee Report"

            # Headers - ensure these match the order from get_all_employees
            headers = ["ID", "Name", "Position", "Hourly Rate", "Hire Date", "Date of Birth",
                       "Address", "Phone", "Email", "Emergency Contact", "Tax ID",
                       "Bank Name", "Account Number", "Routing Number"]
            sheet.append(headers)

            # Apply header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, cell in enumerate(sheet[1]):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx + 1)].width = 15 # Default width

            # Data rows
            for row_idx, emp in enumerate(employee_data):
                # emp: (id, name, position, hourly_rate, hire_date, dob, address, phone, email, emergency_contact, tax_id, bank_name, account_number, routing_number)
                sheet.append(list(emp)) # Append all values directly
                # Format hourly rate column
                sheet.cell(row=row_idx + 2, column=4).number_format = '"$"#,##0.00' # Hourly Rate is 4th column (index 3)

            # Auto-size columns
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Report Generated", f"Employee report saved to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate employee report: {e}")
            print(f"Error generating employee report: {e}")

    def setup_auto_save(self):
        # Auto-save every 5 minutes (300000 milliseconds)
        self.after(300000, self.auto_save)

    def auto_save(self):
        """
        Automatically saves pending data to the database.
        """
        try:
            payroll_db.save_all_data(self.db_path)
        except:
            pass  # Silent auto-save - don't disturb user with errors
        self.after(300000, self.auto_save)  # Reschedule auto-save

    def on_closing(self):
        """
        Handles window closing events to ensure data is saved.
        """
        if messagebox.askokcancel("Quit Professional Payroll System", "Are you sure you want to exit?\n\nAll data will be automatically saved."):
            try:
                payroll_db.save_all_data(self.db_path)
            except:
                pass  # Silent save
            self.destroy()

class SplashScreen:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Professional Payroll System")
        self.root.geometry("450x250")
        self.root.configure(bg='#2E3B4E')
        self.root.resizable(False, False)
        
        # Center the splash screen
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (450 // 2)
        y = (self.root.winfo_screenheight() // 2) - (250 // 2)
        self.root.geometry(f"450x250+{x}+{y}")
        
        # Remove window decorations
        self.root.overrideredirect(True)
        
        # Create main frame
        main_frame = tk.Frame(self.root, bg='#2E3B4E', padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Logo/Title
        title_label = tk.Label(main_frame, text="Professional Payroll System", 
                              font=("Arial", 18, "bold"), fg='white', bg='#2E3B4E')
        title_label.pack(pady=(20, 10))
        
        subtitle_label = tk.Label(main_frame, text="Advanced Employee & Payroll Management", 
                                 font=("Arial", 11), fg='#B0C4DE', bg='#2E3B4E')
        subtitle_label.pack(pady=(0, 30))
        
        # Progress frame
        progress_frame = tk.Frame(main_frame, bg='#2E3B4E')
        progress_frame.pack(fill=tk.X, pady=20)
        
        # Status label
        self.status_label = tk.Label(progress_frame, text="Loading...", 
                                    font=("Arial", 10), fg='white', bg='#2E3B4E')
        self.status_label.pack()
        
        # Simple progress bar simulation
        self.progress_canvas = tk.Canvas(progress_frame, height=6, bg='#1E2B3E', highlightthickness=0)
        self.progress_canvas.pack(fill=tk.X, pady=10)
        
        # Version info
        version_label = tk.Label(main_frame, text="Version 2.0 - Professional Edition", 
                                font=("Arial", 8), fg='#708090', bg='#2E3B4E')
        version_label.pack(side=tk.BOTTOM, pady=(20, 0))
        
        self.progress_value = 0
        self.root.update()
    
    def show_message(self, message):
        self.status_label.config(text=message)
        self.progress_value += 20
        if self.progress_value <= 100:
            self.progress_canvas.delete("all")
            width = self.progress_canvas.winfo_width()
            progress_width = (width * self.progress_value) // 100
            self.progress_canvas.create_rectangle(0, 0, progress_width, 6, fill='#4CAF50', outline='')
        self.root.update()
        # Small delay for visual effect
        import time
        time.sleep(0.3)
    
    def destroy(self):
        self.root.destroy()

if __name__ == "__main__":
    # Create and show splash screen
    splash = SplashScreen()
    splash.show_message("Initializing Professional Payroll System...")
    
    # Get user profile
    splash.show_message("Loading user profiles...")
    profile_name, db_path = select_user_profile()
    if not profile_name:
        splash.destroy()
        exit()
    
    splash.show_message("Setting up database connection...")
    
    # Create main app (hidden initially)
    app = MainApp(profile_name, db_path)
    app.withdraw()  # Hide initially
    
    splash.show_message("Loading application interface...")
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    splash.show_message("Starting Professional Payroll System...")
    
    # Close splash and show main app
    splash.destroy()
    app.deiconify()  # Show the main window
    
    app.mainloop()
