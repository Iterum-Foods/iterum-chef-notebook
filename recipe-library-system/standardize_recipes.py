#!/usr/bin/env python3
"""
Recipe Standardizer & Converter
Converts all recipes to uniform Iterum format for costing and upload
"""

import sys
import os
from pathlib import Path
from datetime import datetime
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

class IterumRecipeConverter:
    """Convert recipes to standardized Iterum format for costing."""
    
    def __init__(self, library_path="recipe_library", output_dir="converted_iterum"):
        self.library_path = Path(library_path)
        self.output_dir = Path(output_dir)
        self.db_path = self.library_path / "recipe_library.db"
        self.output_dir.mkdir(exist_ok=True)
        
    def create_iterum_template(self):
        """Create a new workbook with Iterum format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Recipe"
        
        # Styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        
        return wb, ws
    
    def fill_header_section(self, ws, recipe_data):
        """Fill the header section with recipe metadata."""
        # Row 1: Empty
        # Row 2: Date field
        ws['G2'] = 'Date:'
        ws['H2'] = datetime.now().strftime('%Y-%m-%d')
        
        # Row 3: Recipe name
        ws['A3'] = 'Recipe name:'
        ws['B3'] = recipe_data.get('title', 'Untitled Recipe')
        ws['B3'].font = Font(bold=True, size=14)
        
        # Row 4: Concept
        ws['A4'] = 'Concept:'
        ws['B4'] = recipe_data.get('concept', '')
        ws['G4'] = 'Cuisine:'
        ws['H4'] = recipe_data.get('cuisine', 'unknown')
        
        # Row 5: Submitted by
        ws['A5'] = 'Submitted by:'
        ws['B5'] = recipe_data.get('submitted_by', 'Chef')
        ws['G5'] = 'Category:'
        ws['H5'] = recipe_data.get('category', 'recipe')
        
        # Row 6: Number of Portions
        ws['A6'] = 'Number of Portions: 1 or 24'
        ws['B6'] = recipe_data.get('servings', 1)
        ws['E6'] = 'Operation:'
        
        # Row 7: Serving Size
        ws['A7'] = 'Serving Size Per Person:'
        ws['E7'] = 'Region:'
        
        # Row 8: Cost per portion formula
        ws['A8'] = '=I9/B6'  # Will be calculated
        ws['B8'] = 1
        ws['G8'] = 'Cost per Portion:'
        
        # Row 9: Projected FC%
        ws['A9'] = 'Projected FC%'
        ws['G9'] = 'Cost per Recipe:'
        ws['I9'] = 0  # Will be sum of ingredients
        
        # Row 10: Explanation text
        ws['A10'] = 'EP = Edible Portion    AP$ = As Purchased cost   EP$ = Edible Portion Cost    Unit = unit of measure for ingredients'
        ws['A10'].font = Font(size=9, italic=True)
        
        # Row 11: Yield tools reference
        ws['A11'] = 'For common Yields use these tools:'
        ws['C11'] = 'Produce Yields'
        ws['E11'] = 'Fruit Yields'
        ws['G11'] = 'Fish Yields'
        ws['H11'] = 'Spice Yields'
        
        return ws
    
    def fill_ingredients_section(self, ws, ingredients_data, start_row=13):
        """Fill the ingredients section with table format."""
        # Headers (row 13-14)
        ws[f'B{start_row}'] = 'Recipe Quantity (EP)'
        ws[f'E{start_row}'] = 'Costing'
        ws[f'I{start_row}'] = 'Total'
        
        header_row = start_row + 1
        headers = ['Ingredients', 'Quantity', 'Weight', 'Volume', 'AP$ / Unit', 'Unit', 'Yield %', 'EP$ / Unit', 'Cost']
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Fill ingredient rows
        current_row = header_row + 1
        for ingredient in ingredients_data:
            ws[f'A{current_row}'] = ingredient.get('name', '')
            ws[f'B{current_row}'] = ingredient.get('quantity', '')
            ws[f'C{current_row}'] = ingredient.get('weight', '')
            ws[f'D{current_row}'] = ingredient.get('volume', '')
            ws[f'E{current_row}'] = ingredient.get('ap_cost', '')
            ws[f'F{current_row}'] = ingredient.get('unit', '')
            ws[f'G{current_row}'] = ingredient.get('yield_pct', '')
            ws[f'H{current_row}'] = ingredient.get('ep_cost', '')
            ws[f'I{current_row}'] = ingredient.get('total_cost', '')
            
            current_row += 1
        
        return current_row
    
    def fill_method_section(self, ws, method_text, start_row):
        """Fill the method/instructions section."""
        ws[f'A{start_row}'] = 'Method:'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        current_row = start_row + 1
        
        # Parse method into steps
        if method_text:
            steps = method_text.split('\n')
            for i, step in enumerate(steps, 1):
                if step.strip():
                    # Number the step if not already numbered
                    if not step.strip()[0].isdigit():
                        step = f"{i}. {step.strip()}"
                    ws[f'A{current_row}'] = step
                    current_row += 1
        else:
            ws[f'A{current_row}'] = '1. [TO BE FILLED]'
            current_row += 1
        
        return current_row
    
    def extract_from_existing_excel(self, file_path):
        """Extract data from existing Excel recipe files."""
        try:
            df = pd.read_excel(file_path, header=None)
            
            recipe_data = {}
            ingredients_data = []
            method_text = ""
            
            # Try to find recipe name
            for idx, row in df.iterrows():
                if pd.notna(row[0]):
                    val = str(row[0]).lower()
                    if 'recipe name' in val:
                        recipe_data['title'] = row[1] if pd.notna(row[1]) else Path(file_path).stem
                    elif 'cuisine' in val:
                        recipe_data['cuisine'] = row[1] if pd.notna(row[1]) else 'unknown'
                    elif 'category' in val:
                        recipe_data['category'] = row[1] if pd.notna(row[1]) else 'recipe'
                    elif 'submitted by' in val:
                        recipe_data['submitted_by'] = row[1] if pd.notna(row[1]) else 'Chef'
                    elif 'number of portions' in val:
                        try:
                            recipe_data['servings'] = int(row[1]) if pd.notna(row[1]) else 1
                        except:
                            recipe_data['servings'] = 1
                    elif 'method' in val or 'instructions' in val:
                        # Start collecting method from next row
                        method_start = idx + 1
                        method_lines = []
                        for m_idx in range(method_start, min(method_start + 20, len(df))):
                            if m_idx < len(df) and pd.notna(df.iloc[m_idx, 0]):
                                method_lines.append(str(df.iloc[m_idx, 0]))
                        method_text = '\n'.join(method_lines)
            
            # Try to find ingredients table
            for idx, row in df.iterrows():
                if pd.notna(row[0]) and str(row[0]).lower() == 'ingredients':
                    # Found ingredients header, read next rows
                    for ing_idx in range(idx + 1, min(idx + 30, len(df))):
                        ing_row = df.iloc[ing_idx]
                        if pd.notna(ing_row[0]) and str(ing_row[0]).lower() not in ['', 'method', 'instructions']:
                            ingredient = {
                                'name': ing_row[0],
                                'quantity': ing_row[1] if pd.notna(ing_row[1]) else '',
                                'weight': ing_row[2] if pd.notna(ing_row[2]) else '',
                                'volume': ing_row[3] if pd.notna(ing_row[3]) else '',
                                'ap_cost': ing_row[4] if len(ing_row) > 4 and pd.notna(ing_row[4]) else '',
                                'unit': ing_row[5] if len(ing_row) > 5 and pd.notna(ing_row[5]) else '',
                                'yield_pct': ing_row[6] if len(ing_row) > 6 and pd.notna(ing_row[6]) else '',
                                'ep_cost': ing_row[7] if len(ing_row) > 7 and pd.notna(ing_row[7]) else '',
                                'total_cost': ing_row[8] if len(ing_row) > 8 and pd.notna(ing_row[8]) else '',
                            }
                            ingredients_data.append(ingredient)
                        else:
                            break
                    break
            
            # If no structured data found, create placeholder
            if not recipe_data.get('title'):
                recipe_data['title'] = Path(file_path).stem
            
            if not ingredients_data:
                ingredients_data = [
                    {'name': '[Ingredient 1]', 'quantity': '', 'weight': '', 'volume': '', 
                     'ap_cost': '', 'unit': '', 'yield_pct': '', 'ep_cost': '', 'total_cost': ''},
                ]
            
            return recipe_data, ingredients_data, method_text
            
        except Exception as e:
            print(f"   Warning: Could not fully parse {Path(file_path).name}: {e}")
            return {
                'title': Path(file_path).stem,
                'cuisine': 'unknown',
                'category': 'recipe'
            }, [{'name': '[TO BE FILLED]', 'quantity': '', 'weight': '', 'volume': '', 
                 'ap_cost': '', 'unit': '', 'yield_pct': '', 'ep_cost': '', 'total_cost': ''}], ""
    
    def convert_recipe(self, file_path, recipe_metadata=None):
        """Convert a single recipe to Iterum format."""
        print(f"   Converting: {Path(file_path).name}")
        
        # Extract data from file
        recipe_data, ingredients_data, method_text = self.extract_from_existing_excel(file_path)
        
        # Merge with metadata from database if available
        if recipe_metadata:
            recipe_data['cuisine'] = recipe_metadata.get('cuisine_type', recipe_data.get('cuisine', 'unknown'))
            recipe_data['category'] = recipe_metadata.get('category', recipe_data.get('category', 'recipe'))
            recipe_data['title'] = recipe_metadata.get('title', recipe_data.get('title'))
        
        # Create new workbook with template
        wb, ws = self.create_iterum_template()
        
        # Fill sections
        self.fill_header_section(ws, recipe_data)
        ing_end_row = self.fill_ingredients_section(ws, ingredients_data)
        method_end_row = self.fill_method_section(ws, method_text, ing_end_row + 3)
        
        # Add copyright footer
        ws[f'A{method_end_row + 10}'] = 'This file cannot be used or sold as a means of generating revenue without prior permission in writing.'
        ws[f'A{method_end_row + 10}'].font = Font(size=8, italic=True)
        
        # Save
        output_filename = f"{recipe_data['title']}.xlsx"
        # Sanitize filename
        output_filename = "".join(c for c in output_filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
        output_path = self.output_dir / output_filename
        
        wb.save(output_path)
        print(f"   [OK] Saved: {output_filename}")
        
        return output_path
    
    def convert_all_recipes(self):
        """Convert all recipes in the library to Iterum format."""
        print("=" * 80)
        print("           RECIPE STANDARDIZER & CONVERTER")
        print("           Converting to Iterum Format for Costing")
        print("=" * 80)
        
        # Get all recipes from database
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT id, title, cuisine_type, category, library_path FROM recipes")
        recipes = cursor.fetchall()
        conn.close()
        
        print(f"\nFound {len(recipes)} recipes to convert\n")
        
        converted = []
        errors = []
        
        for recipe_id, title, cuisine, category, lib_path in recipes:
            try:
                metadata = {
                    'title': title,
                    'cuisine_type': cuisine,
                    'category': category
                }
                output_path = self.convert_recipe(lib_path, metadata)
                converted.append(output_path)
            except Exception as e:
                print(f"   [ERROR] Error converting {title}: {e}")
                errors.append((title, str(e)))
        
        # Summary
        print("\n" + "=" * 80)
        print(f"[OK] Successfully converted: {len(converted)} recipes")
        if errors:
            print(f"[ERROR] Errors: {len(errors)} recipes")
            for title, error in errors:
                print(f"   - {title}: {error}")
        
        print(f"\nOutput directory: {self.output_dir.absolute()}")
        print("=" * 80)
        
        return converted, errors

def main():
    print("\n" + "=" * 80)
    print("           RECIPE STANDARDIZER FOR ITERUM")
    print("=" * 80)
    print("\nThis tool converts ALL recipes to the standardized Iterum format:")
    print("  - Professional costing layout")
    print("  - Ingredient quantity tracking")
    print("  - AP/EP cost fields")
    print("  - Yield percentage tracking")
    print("  - Ready for upload to Iterum Chef's Notebook")
    print("\n" + "=" * 80)
    
    # Check if running non-interactively
    if len(sys.argv) > 1 and sys.argv[1] == '--auto':
        print("\nRunning in automatic mode...")
    else:
        try:
            input("\nPress Enter to begin conversion...")
        except EOFError:
            print("\nStarting conversion...")
    
    converter = IterumRecipeConverter()
    converted, errors = converter.convert_all_recipes()
    
    if converted:
        print("\n" + "=" * 80)
        print("CONVERSION COMPLETE!")
        print("=" * 80)
        print(f"\nConverted recipes are in: {converter.output_dir.absolute()}")
        print("\nNext steps:")
        print("  1. Open the converted Excel files")
        print("  2. Fill in any missing costing data (AP$, Units, Yields)")
        print("  3. Upload to Iterum Chef's Notebook")
        print("  4. Use for recipe costing and menu planning")

if __name__ == "__main__":
    main()

