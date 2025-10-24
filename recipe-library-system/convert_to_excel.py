import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# Optional: import docx, PyPDF2, etc. for advanced extraction
# from docx import Document
# import PyPDF2

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Recipe_Template.xlsx')
RECIPE_DIR = os.path.join(os.path.dirname(__file__), 'recipe_library')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'converted_excel')

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Fields in the template (update if template changes)
TEMPLATE_FIELDS = [
    'Type', 'Category', 'Title', 'Restaurant', 'Cuisine', 'Difficulty', 'Tags',
    'Servings', 'Cooking Time', 'Source File Name', 'Notes'
]
INGREDIENTS_HEADER = ['Ingredient Name', 'Amount', 'Unit']

# Helper: Extract info from a file (stub for now)
def extract_recipe_info(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    info = {field: 'TO FILL' for field in TEMPLATE_FIELDS}
    ingredients = [['TO FILL', 'TO FILL', 'TO FILL']]
    instructions = 'TO FILL'
    # TODO: Implement extraction for .xlsx, .docx, .pdf, .txt
    # For now, just use filename as Title
    info['Title'] = os.path.splitext(os.path.basename(filepath))[0]
    info['Source File Name'] = os.path.basename(filepath)
    return info, ingredients, instructions

# Helper: Write to Excel using template
def write_to_excel(info, ingredients, instructions, output_path):
    if os.path.exists(TEMPLATE_PATH):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        for i, field in enumerate(TEMPLATE_FIELDS, 1):
            ws[f'{get_column_letter(i)}1'] = field
        # Ingredients and instructions headers
        ws.append([])
        ws.append(INGREDIENTS_HEADER)
    # Fill in main fields
    for i, field in enumerate(TEMPLATE_FIELDS, 1):
        ws[f'{get_column_letter(i)}2'] = info.get(field, 'TO FILL')
    # Ingredients table (start at row 4)
    row = 4
    for ing in ingredients:
        for j, val in enumerate(ing, 1):
            ws[f'{get_column_letter(j)}{row}'] = val
        row += 1
    # Instructions (put after ingredients)
    ws[f'A{row+1}'] = 'Instructions'
    ws[f'B{row+1}'] = instructions
    wb.save(output_path)

# Main loop
for fname in os.listdir(RECIPE_DIR):
    fpath = os.path.join(RECIPE_DIR, fname)
    if not os.path.isfile(fpath):
        continue
    ext = os.path.splitext(fname)[1].lower()
    if ext not in ['.xlsx', '.xls', '.csv', '.docx', '.doc', '.pdf', '.txt', '.md', '.html', '.htm', '.json']:
        continue
    info, ingredients, instructions = extract_recipe_info(fpath)
    outname = os.path.splitext(fname)[0] + '_converted.xlsx'
    outpath = os.path.join(OUTPUT_DIR, outname)
    write_to_excel(info, ingredients, instructions, outpath)
    print(f'Converted: {fname} -> {outname}') 